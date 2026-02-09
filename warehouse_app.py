import streamlit as st
import sqlite3
import pandas as pd
from io import BytesIO
from datetime import date
import re
import os
import logging

# Лог-файл в папке с приложением
LOG_PATH = os.path.join(os.path.dirname(__file__), "fifo_recalc.log")
logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format='%(asctime)s %(levelname)s: %(message)s',
)
logger = logging.getLogger("warehouse_app")
logger.info("Starting warehouse_app, log file: %s", LOG_PATH)


# --- session state (чтобы ввод номера/даты не сбрасывался при rerun) ---
if "mp_month_sale_date" not in st.session_state:
    st.session_state.mp_month_sale_date = date.today()
if "mp_month_sale_number" not in st.session_state:
    st.session_state.mp_month_sale_number = ""
if "mp_month_overwrite" not in st.session_state:
    st.session_state.mp_month_overwrite = False



def normalize_order_number(val):
    """Нормализует № заказа для сопоставления между разными файлами."""
    if pd.isna(val):
        return ""
    s = str(val)
    # убрать .0 у числовых значений
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    # убрать пробелы/неразрывные/zero-width
    s = re.sub(r"\s+", "", s)
    s = s.replace("\u200b", "").replace("\ufeff", "")
    return s.strip().upper()


def order_match_key(val):
    """Ключ сопоставления заказа между WB и Европочтой.

    Пытаемся достать самый длинный блок цифр (часто именно он совпадает),
    иначе оставляем нормализованную буквенно-цифровую строку.
    """
    if pd.isna(val):
        return ""
    s = str(val)
    # убрать .0 у числовых значений
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    s = s.replace("\u200b", "").replace("\ufeff", "")
    s = re.sub(r"\s+", "", s).strip().upper()
    # оставить только латиницу/цифры (убрать дефисы, слэши и т.п.)
    s_alnum = re.sub(r"[^0-9A-Z]", "", s)
    # самый длинный блок цифр
    digits = re.findall(r"\d+", s_alnum)
    if digits:
        longest = max(digits, key=len)
        if len(longest) >= 6:
            return longest
    return s_alnum



def _norm_col(s: str) -> str:
    """Нормализация заголовков колонок для устойчивого сопоставления."""
    s = str(s).replace("\n", " ").replace("\r", " ")
    s = s.strip().lower()
    s = s.replace("ё", "е")
    s = s.replace("№", "n")
    s = re.sub(r"\s+", " ", s)
    s = s.replace("руб.", "руб").replace("шт.", "шт")
    return s

def standardize_sales_columns(df: pd.DataFrame):
    """Пытается привести заголовки файла продаж к каноническим.
    Возвращает (df2, missing_cols, rename_map).
    """
    if df is None or df.empty:
        return df, ["файл пустой"], {}

    # нормализованное_имя -> оригинальное
    norm_to_orig = {_norm_col(c): c for c in df.columns}

    aliases = {
        "№ заказа": ["n заказа", "номер заказа", "id заказа", "заказ n", "номер отправления", "nзаказа"],
        "ШК": ["шк", "штрихкод", "шк товара", "barcode", "ean", "ean13"],
        "Наименование товара": ["наименование товара", "товар", "наименование", "название товара", "предмет"],
        "Количество в заказе, шт.": ["количество в заказе, шт", "количество в заказе", "количество, шт", "кол-во, шт", "количество", "кол-во"],
        "Цена продажи, руб.": ["цена продажи, руб", "цена продажи", "цена, руб", "цена"],
        "К оплате с НДС, руб.": ["к оплате с ндс, руб", "к оплате с ндс", "к оплате, руб", "итого, руб", "итого к оплате, руб", "сумма к оплате, руб"],
    }

    rename_map = {}

    # 1) точные совпадения по нормализации
    for canonical, vars_ in aliases.items():
        found = None
        for v in [canonical] + vars_:
            key = _norm_col(v)
            if key in norm_to_orig:
                found = norm_to_orig[key]
                break

        # 2) подстрочный поиск (на случай "ШК товара (EAN)" и т.п.)
        if found is None:
            for norm_name, orig in norm_to_orig.items():
                if canonical == "ШК" and ("шк" in norm_name or "штрихкод" in norm_name or "ean" in norm_name or "barcode" in norm_name):
                    found = orig
                    break
                if canonical == "№ заказа" and ("заказ" in norm_name and ("n" in norm_name or "номер" in norm_name or "id" in norm_name or "отправлен" in norm_name)):
                    found = orig
                    break
                if canonical == "Наименование товара" and ("наимен" in norm_name or "назван" in norm_name or "товар" in norm_name):
                    found = orig
                    break
                if canonical == "Количество в заказе, шт." and ("кол" in norm_name and ("шт" in norm_name or "кол-во" in norm_name or "количество" in norm_name)):
                    found = orig
                    break
                if canonical == "Цена продажи, руб." and ("цен" in norm_name and ("прод" in norm_name or "продаж" in norm_name)):
                    found = orig
                    break
                if canonical == "К оплате с НДС, руб." and ("оплат" in norm_name or "итого" in norm_name or "к оплате" in norm_name) and ("ндс" in norm_name or "с ндс" in norm_name):
                    found = orig
                    break

        if found is not None:
            rename_map[found] = canonical

    df2 = df.rename(columns=rename_map)

    required = [
        "№ заказа",
        "ШК",
        "Наименование товара",
        "Количество в заказе, шт.",
        "Цена продажи, руб.",
        "К оплате с НДС, руб.",
    ]
    missing = [c for c in required if c not in df2.columns]
    return df2, missing, rename_map

# ===================================
# БАЗА ДАННЫХ
# ===================================
conn = sqlite3.connect("warehouse.db", check_same_thread=False)
c = conn.cursor()

# Таблицы прихода
c.execute("""
CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    article TEXT UNIQUE NOT NULL,
    name TEXT NOT NULL,
    unit TEXT
)
""")

c.execute("""
CREATE TABLE IF NOT EXISTS invoices (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_number TEXT NOT NULL,
    invoice_date TEXT NOT NULL,
    supplier TEXT
)
""")

c.execute("""
CREATE TABLE IF NOT EXISTS invoice_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_id INTEGER NOT NULL,
    product_id INTEGER NOT NULL,
    qty REAL NOT NULL,
    price REAL NOT NULL,
    vat_percent REAL NOT NULL,
    total REAL NOT NULL,
    total_with_vat REAL NOT NULL,
    FOREIGN KEY (invoice_id) REFERENCES invoices(id),
    FOREIGN KEY (product_id) REFERENCES products(id)
)
""")

# Таблицы продаж
c.execute("""
CREATE TABLE IF NOT EXISTS sales (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sale_number TEXT NOT NULL,
    sale_date TEXT NOT NULL,
    comment TEXT
)
""")

c.execute("""
CREATE TABLE IF NOT EXISTS sale_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sale_id INTEGER NOT NULL,
    product_id INTEGER NOT NULL,
    qty REAL NOT NULL,
    price REAL NOT NULL,
    total REAL NOT NULL,
    FOREIGN KEY (sale_id) REFERENCES sales(id),
    FOREIGN KEY (product_id) REFERENCES products(id)
)
""")
conn.commit()

# --- ensure optional products.barcode column (для сопоставления штрихкодов) ---
try:
    cols = [r[1] for r in c.execute("PRAGMA table_info(products)").fetchall()]
    if "barcode" not in cols:
        c.execute("ALTER TABLE products ADD COLUMN barcode TEXT")
        conn.commit()
except Exception:
    # если БД read-only или другая проблема — просто игнорируем; будем использовать article как fallback
    pass


# --- ensure optional sale_items columns for marketplace accounting ---
try:
    si_cols = [r[1] for r in c.execute("PRAGMA table_info(sale_items)").fetchall()]
    to_add = []
    if "gross_price" not in si_cols:
        to_add.append(("gross_price", "REAL"))
    if "gross_total" not in si_cols:
        to_add.append(("gross_total", "REAL"))
    if "mp_fee" not in si_cols:
        to_add.append(("mp_fee", "REAL"))
    if "mp_delivery" not in si_cols:
        to_add.append(("mp_delivery", "REAL"))
    if "net_total" not in si_cols:
        to_add.append(("net_total", "REAL"))
    for col, ctype in to_add:
        c.execute(f"ALTER TABLE sale_items ADD COLUMN {col} {ctype}")
    if to_add:
        conn.commit()
except Exception:
    # если БД read-only или другая проблема — просто игнорируем
    pass


# ===================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ===================================
# --- Товары ---

def get_products(search_text=""):
    if search_text:
        return pd.read_sql(
            "SELECT id, article, name, unit FROM products WHERE article LIKE ? OR name LIKE ? ORDER BY name",
            conn, params=(f"%{search_text}%", f"%{search_text}%"))
    return pd.read_sql("SELECT id, article, name, unit FROM products ORDER BY name", conn)

def insert_product(article, name, unit):
    c.execute("INSERT OR IGNORE INTO products (article, name, unit) VALUES (?, ?, ?)", (article, name, unit))
    conn.commit()

# Создаём таблицу sale_fifo при старте приложения
try:
    ensure_sale_fifo(conn)
except Exception:
    pass

# --- Накладные ---

def get_invoices_df():
    return pd.read_sql("SELECT id, invoice_number, invoice_date, supplier FROM invoices ORDER BY invoice_date DESC", conn)

def insert_invoice(number, date_str, supplier):
    c.execute("INSERT INTO invoices (invoice_number, invoice_date, supplier) VALUES (?, ?, ?)", (number, date_str, supplier))
    conn.commit()
    return c.lastrowid

def add_item(invoice_id, product_id, qty, price, vat_percent):
    total = qty * price
    total_with_vat = total * (1 + vat_percent / 100)
    c.execute("""
        INSERT INTO invoice_items (invoice_id, product_id, qty, price, vat_percent, total, total_with_vat)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (invoice_id, product_id, qty, price, vat_percent, total, total_with_vat))
    conn.commit()

def get_invoice_items_df(invoice_id):
    return pd.read_sql("""
        SELECT ii.id, p.article AS Артикул, p.name AS Наименование, p.unit AS Ед, 
               ii.qty AS Количество, ii.price AS Цена, ii.vat_percent AS НДС, 
               ii.total AS Сумма, ii.total_with_vat AS Сумма_с_НДС
        FROM invoice_items ii
        JOIN products p ON ii.product_id = p.id
        WHERE ii.invoice_id = ?
        ORDER BY ii.id
    """, conn, params=(invoice_id,))

def update_item(item_id, qty, price, vat_percent):
    total = qty * price
    total_with_vat = total * (1 + vat_percent / 100)
    c.execute("""
        UPDATE invoice_items
        SET qty=?, price=?, vat_percent=?, total=?, total_with_vat=?
        WHERE id=?
    """, (qty, price, vat_percent, total, total_with_vat, item_id))
    conn.commit()

def delete_item(item_id):
    c.execute("DELETE FROM invoice_items WHERE id=?", (item_id,))
    conn.commit()

def get_invoices_df():
    return pd.read_sql("SELECT id, invoice_number, invoice_date, supplier FROM invoices ORDER BY invoice_date DESC", conn)

def get_invoice_sum(invoice_id):
    s = pd.read_sql("SELECT SUM(total_with_vat) AS total FROM invoice_items WHERE invoice_id=?", conn, params=(invoice_id,))
    return round(s.iloc[0]["total"] or 0, 2)

def insert_invoice(number, date_str, supplier):
    c.execute("INSERT INTO invoices (invoice_number, invoice_date, supplier) VALUES (?, ?, ?)", (number, date_str, supplier))
    conn.commit()
    return c.lastrowid

def get_invoice_items_df(invoice_id):
    return pd.read_sql("""
        SELECT ii.id, p.article AS Артикул, p.name AS Наименование, p.unit AS Ед, 
               ii.qty AS Количество, ii.price AS Цена, ii.vat_percent AS НДС, 
               ii.total AS Сумма, ii.total_with_vat AS Сумма_с_НДС
        FROM invoice_items ii
        JOIN products p ON ii.product_id = p.id
        WHERE ii.invoice_id = ?
        ORDER BY ii.id
    """, conn, params=(invoice_id,))

def update_invoice_item(item_id, qty, price, vat_percent):
    total = qty * price
    total_with_vat = total * (1 + vat_percent / 100)
    c.execute("""
        UPDATE invoice_items
        SET qty=?, price=?, vat_percent=?, total=?, total_with_vat=?
        WHERE id=?
    """, (qty, price, vat_percent, total, total_with_vat, item_id))
    conn.commit()

def delete_invoice_item(item_id):
    c.execute("DELETE FROM invoice_items WHERE id=?", (item_id,))
    conn.commit()

def invoice_totals(invoice_id):
    df = pd.read_sql("""
        SELECT SUM(qty) AS qty_sum, SUM(total) AS total_sum, SUM(total_with_vat) AS total_w_vat
        FROM invoice_items WHERE invoice_id=?
    """, conn, params=(invoice_id,))
    if df.empty:
        return 0, 0, 0
    row = df.iloc[0]
    return (row["qty_sum"] or 0, row["total_sum"] or 0, row["total_w_vat"] or 0)

# --- Продажи ---

def add_sale_item(sale_id, product_id, qty, price):
    total = qty * price
    c.execute("""
        INSERT INTO sale_items (sale_id, product_id, qty, price, total)
        VALUES (?, ?, ?, ?, ?)
    """, (sale_id, product_id, qty, price, total))
    conn.commit()
    try:
        recalc_fifo_for_sale(conn, sale_id)
    except Exception:
        pass


def add_sale_item_mp(sale_id, product_id, qty, price, gross_total, mp_fee, mp_delivery, net_total):
    """Добавляет строку продажи с раздельным учётом комиссии/доставки (для маркетплейса).
    Если в БД нет расширенных колонок — тихо падаем обратно на обычную вставку.
    """
    total = float(gross_total)  # total в базе = грязная сумма по строке
    gross_price = float(price) if qty else float(price)
    try:
        c.execute("""
            INSERT INTO sale_items (
                sale_id, product_id, qty, price, total,
                gross_price, gross_total, mp_fee, mp_delivery, net_total
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            sale_id, product_id, float(qty), float(price), float(total),
            float(gross_price), float(gross_total), float(mp_fee), float(mp_delivery), float(net_total)
        ))
        conn.commit()
    except Exception:
        # fallback для старых БД
        add_sale_item(sale_id, product_id, qty, price)
    else:
        try:
            recalc_fifo_for_sale(conn, sale_id)
        except Exception:
            pass


def get_sales_df():
    df = pd.read_sql("SELECT id, sale_number, sale_date, comment FROM sales ORDER BY sale_date DESC", conn)
    if not df.empty:
        sums = pd.read_sql("SELECT sale_id, SUM(total) AS total FROM sale_items GROUP BY sale_id", conn)
        df = df.merge(sums, left_on="id", right_on="sale_id", how="left").fillna(0)
        df["Сумма чека"] = df["total"].round(2)
        df.drop(columns=["total", "sale_id"], inplace=True, errors="ignore")
        df = df[["id", "sale_number", "sale_date", "Сумма чека", "comment"]]
    return df

def insert_sale(number, date_str, comment=""):
    c.execute("INSERT INTO sales (sale_number, sale_date, comment) VALUES (?, ?, ?)", (number, date_str, comment))
    conn.commit()
    return c.lastrowid

def get_sale_items_df(sale_id):
    return pd.read_sql("""
        SELECT si.id, p.article AS Артикул, p.name AS Наименование, p.unit AS Ед,
               si.qty AS Количество, si.price AS Цена, si.total AS Сумма
        FROM sale_items si
        JOIN products p ON si.product_id = p.id
        WHERE si.sale_id = ?
        ORDER BY si.id
    """, conn, params=(sale_id,))

def update_sale_item(item_id, qty, price):
    total = qty * price
    c.execute("UPDATE sale_items SET qty=?, price=?, total=? WHERE id=?", (qty, price, total, item_id))
    conn.commit()
    try:
        row = c.execute("SELECT sale_id FROM sale_items WHERE id = ?", (item_id,)).fetchone()
        if row:
            recalc_fifo_for_sale(conn, int(row[0]))
    except Exception:
        pass

def delete_sale_item(item_id):
    try:
        row = c.execute("SELECT sale_id FROM sale_items WHERE id = ?", (item_id,)).fetchone()
        sale_id = int(row[0]) if row else None
    except Exception:
        sale_id = None
    c.execute("DELETE FROM sale_items WHERE id=?", (item_id,))
    conn.commit()
    try:
        if sale_id is not None:
            recalc_fifo_for_sale(conn, sale_id)
    except Exception:
        pass

def sale_totals(sale_id):
    df = pd.read_sql("SELECT SUM(qty) AS qty_sum, SUM(total) AS total_sum FROM sale_items WHERE sale_id=?", conn, params=(sale_id,))
    if df.empty:
        return 0, 0
    row = df.iloc[0]
    return (row["qty_sum"] or 0, row["total_sum"] or 0)

# --- ФУНКЦИИ ДЛЯ ОТЧЁТА "ОСТАТКИ НА ДАТУ ---

def get_stock_on_date(selected_date):
    """Возвращает остатки товаров на указанную дату."""
    # Подсчёт приходов
    df_in = pd.read_sql("""
        SELECT 
            p.id AS product_id,
            p.article AS Артикул,
            p.name AS Наименование,
            p.unit AS Ед,
            SUM(ii.qty) AS Приход,
            AVG(ii.price) AS Закуп_цена
        FROM invoice_items ii
        JOIN invoices i ON i.id = ii.invoice_id
        JOIN products p ON p.id = ii.product_id
        WHERE DATE(i.invoice_date) <= DATE(?)
        GROUP BY p.id, p.article, p.name, p.unit
    """, conn, params=(str(selected_date),))

    # Подсчёт расходов
    df_out = pd.read_sql("""
        SELECT p.id AS product_id, SUM(si.qty) AS Расход
        FROM sale_items si
        JOIN sales s ON s.id = si.sale_id
        JOIN products p ON p.id = si.product_id
        WHERE DATE(s.sale_date) <= DATE(?)
        GROUP BY p.id
    """, conn, params=(str(selected_date),))

    # Объединяем и считаем остаток
    df = pd.merge(df_in, df_out, on="product_id", how="left").fillna(0)
    df["Остаток"] = df["Приход"] - df["Расход"]
    df["Сумма"] = (df["Остаток"] * df["Закуп_цена"]).round(2)

    # Убираем нулевые остатки
    df = df[df["Остаток"] > 0].reset_index(drop=True)

    # Добавляем колонку "№" (нумерация с 1)
    df.index = df.index + 1
    df.insert(0, "№", df.index)

    # Упорядочиваем столбцы
    df = df[["№", "Артикул", "Наименование", "Ед", "Остаток", "Закуп_цена", "Сумма"]]

    return df

def export_stock_to_excel(df, selected_date):
    """Создаёт Excel-файл с остатками и возвращает буфер BytesIO."""
    from io import BytesIO
    buffer = BytesIO()
    
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Остатки")
        workbook = writer.book
        worksheet = writer.sheets["Остатки"]

        # Форматирование колонок
        center_fmt = workbook.add_format({"align": "center"})
        money_fmt = workbook.add_format({"num_format": "#,##0.00", "align": "center"})

        worksheet.set_column("A:A", 6, center_fmt)     # №
        worksheet.set_column("B:B", 15, center_fmt)    # Артикул
        worksheet.set_column("C:C", 45)                # Наименование
        worksheet.set_column("D:D", 8, center_fmt)     # Ед
        worksheet.set_column("E:E", 12, money_fmt)     # Остаток
        worksheet.set_column("F:F", 12, money_fmt)     # Закуп_цена
        worksheet.set_column("G:G", 14, money_fmt)     # Сумма

        # Итоги
        last = len(df) + 3
        worksheet.write(f"A{last}", "Итого:")
        worksheet.write(f"E{last}", float(df["Остаток"].sum()), money_fmt)
        worksheet.write(f"G{last}", float(df["Сумма"].sum()), money_fmt)

    buffer.seek(0)
    return buffer


# --- ФУНКЦИИ ДЛЯ "КНИГА УЧЁТА ТОВАРОВ"

import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict

def export_book(selected_year: int):
    """
    Формирует строки для 'Книга учёта товаров' за конкретный год.
    - каждая строка = ОТДЕЛЬНАЯ ПАРТИЯ прихода (позиция накладной)
    - A,B  : поставщик + накладная + дата, наименование товара
    - C,D  : остаток на начало года (для первого года = 0)
    - AU   : цена закупки (приход IV квартал)
    - AV   : количество закупки (приход IV квартал)
    - AW   : стоимость закупки (приход IV квартал)
    - BC   : номера чеков продажи, относящиеся к ЭТОЙ партии (через запятую)
    - BD   : цена продажи (одно значение, если все одинаковые; иначе – через запятую)
    - BE   : количества продаж по чекам (через запятую, в том же порядке, что BC)
    - BF   : расходы = Σ(кол-во списания * закупочная цена этой партии)
    - BG   : остаток количества по партии (после FIFO-списания)
    - BH   : остаток стоимости по партии = BG * закупочная цена партии
    """

    from io import BytesIO

    template_path = "Книга учёта товаров.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active  # Sheet1

def delete_sale(sale_id):
    """Удаляет чек и все его позиции."""
    try:
        c.execute("DELETE FROM sale_fifo WHERE sale_item_id IN (SELECT id FROM sale_items WHERE sale_id=?)", (sale_id,))
    except Exception:
        pass
    c.execute("DELETE FROM sale_items WHERE sale_id=?", (sale_id,))
    c.execute("DELETE FROM sales WHERE id=?", (sale_id,))
    conn.commit()

def get_sale_id_by_number(sale_number: str):
    """Возвращает id чека по номеру (если существует), иначе None."""
    row = c.execute("SELECT id FROM sales WHERE sale_number = ?", (sale_number,)).fetchone()
    return int(row[0]) if row else None


def upsert_marketplace_sale(sale_number: str, sale_date_str: str, comment: str, overwrite: bool):
    """Создаёт чек или обновляет существующий (по номеру). Возвращает sale_id."""
    existing_id = get_sale_id_by_number(sale_number)
    if existing_id is None:
        c.execute(
            "INSERT INTO sales (sale_number, sale_date, comment) VALUES (?, ?, ?)",
            (sale_number, sale_date_str, comment),
        )
        conn.commit()
        return int(c.lastrowid)

    if overwrite:
        # обновим шапку и удалим старые позиции
        c.execute(
            "UPDATE sales SET sale_date=?, comment=? WHERE id=?",
            (sale_date_str, comment, existing_id),
        )
        c.execute("DELETE FROM sale_items WHERE sale_id=?", (existing_id,))
        conn.commit()

    return existing_id


def push_marketplace_sales_to_db(merged_df: pd.DataFrame, prefix: str = "MP", overwrite: bool = False):
    """
    Загружает интернет-продажи из merged_df в обычные таблицы sales / sale_items.
    1 заказ = 1 чек. Номер чека: {prefix}-{order_number}

    Возвращает (created, skipped, updated, missing_products_df, preview_df)
    """
    if merged_df is None or merged_df.empty:
        return 0, 0, 0, pd.DataFrame(), pd.DataFrame()

    # Определяем колонки
    order_col = "№ заказа"
    qty_col = "Количество"
    price_col_candidates = ["Цена продажи, руб.", "Цена продажи", "Цена продажи, руб"]
    fee_col = "Вознаграждение площадки"
    delivery_col = "Стоимость доставки"

    price_col = None
    for cand in price_col_candidates:
        if cand in merged_df.columns:
            price_col = cand
            break
    if price_col is None:
        raise ValueError("Не найдена колонка с ценой продажи в итоговой таблице.")

    # Дата чека: в merged_df она одинаковая (последний день периода), берём первую
    sale_date_val = merged_df.get("Дата продажи", None)
    if sale_date_val is not None:
        try:
            sale_date_str = str(pd.to_datetime(merged_df["Дата продажи"].iloc[0]).date())
        except Exception:
            sale_date_str = str(merged_df["Дата продажи"].iloc[0])
    else:
        # fallback: берём из периодов/текущего
        sale_date_str = str(pd.Timestamp.today().date())

    # Готовим превью по заказам
    orders = (
        merged_df[order_col]
        .dropna()
        .astype(str)
        .apply(normalize_order_number)
        .loc[lambda s: (s != "") & (s != "NAN")]
        .drop_duplicates()
        .sort_values()
        .tolist()
    )

    preview_rows = []
    created = skipped = updated = 0
    missing_products = []

    comment = f"Интернет-торговля ({prefix})"

    # Группируем по заказам
    for order in orders:
        sale_number = f"{prefix}-{order}"
        existing_id = get_sale_id_by_number(sale_number)
        exists = existing_id is not None

        # позиции именно этого заказа
        df_o = merged_df.copy()
        df_o["_order_norm"] = df_o[order_col].astype(str).apply(normalize_order_number)
        df_o = df_o[df_o["_order_norm"] == order].copy()

        # Защита от дублей: внутри одного заказа одинаковый артикул может встретиться несколько раз.
        # Схлопываем в одну строку: qty суммируем, total суммируем, цену считаем как total/qty.
        if "Артикул" in df_o.columns:
            df_o["_article_norm"] = df_o["Артикул"].astype(str).str.strip()
            df_o = df_o[(df_o["_article_norm"].notna()) & (df_o["_article_norm"].str.lower() != "nan") & (df_o["_article_norm"] != "")]
            if not df_o.empty:
                # пересчёт total на всякий случай
                df_o["_qty_num"] = pd.to_numeric(df_o.get(qty_col, 0), errors="coerce").fillna(0.0)
                df_o["_price_num"] = pd.to_numeric(df_o.get(price_col, 0), errors="coerce").fillna(0.0)
                df_o["_total_num"] = pd.to_numeric(df_o.get("Сумма продажи", None), errors="coerce")
                if "_total_num" in df_o.columns and df_o["_total_num"].notna().any():
                    pass
                else:
                    df_o["_total_num"] = (df_o["_qty_num"] * df_o["_price_num"]).round(2)

                grouped = (
                    df_o.groupby("_article_norm", as_index=False)
                    .agg({
                        "Наименование товара": "first",
                        qty_col: "sum",
                        "_total_num": "sum"
                    })
                )
                grouped[price_col] = (grouped["_total_num"] / grouped[qty_col].replace(0, pd.NA)).fillna(0.0).round(4)
                df_o = grouped.rename(columns={"_article_norm": "Артикул"}).copy()
                df_o["Сумма продажи"] = df_o["_total_num"].round(2)
                df_o.drop(columns=["_total_num"], inplace=True, errors="ignore")

        items_cnt = len(df_o)

        # Если чек уже есть и overwrite=False — пропускаем
        if exists and not overwrite:
            skipped += 1
            preview_rows.append({
                "№ заказа": order,
                "Номер чека (sale_number)": sale_number,
                "Дата чека": sale_date_str,
                "Строк в заказе": items_cnt,
                "Статус": "уже есть (пропущен)"
            })
            continue

        sale_id = upsert_marketplace_sale(sale_number, sale_date_str, comment, overwrite=overwrite)
        if exists and overwrite:
            updated += 1
            status = "обновлён"
        else:
            created += 1
            status = "создан"

        # Вставляем позиции
        for _, r in df_o.iterrows():
            article = str(r.get("Артикул", "")).strip()
            if article == "" or article.lower() == "nan":
                continue

            prod = c.execute("SELECT id FROM products WHERE article = ?", (article,)).fetchone()
            if not prod:
                missing_products.append({
                    "№ заказа": order,
                    "Артикул": article,
                    "Наименование товара": r.get("Наименование товара", ""),
                    "Количество": r.get(qty_col, 0),
                    "Цена продажи": r.get(price_col, 0),
                })
                continue

            product_id = int(prod[0])
            qty = float(pd.to_numeric(r.get(qty_col, 0), errors="coerce") or 0)
            price = float(pd.to_numeric(r.get(price_col, 0), errors="coerce") or 0)
            if qty == 0:
                continue

            add_sale_item(sale_id, product_id, qty, price)

        preview_rows.append({
            "№ заказа": order,
            "Номер чека (sale_number)": sale_number,
            "Дата чека": sale_date_str,
            "Строк в заказе": items_cnt,
            "Статус": status
        })

    missing_products_df = pd.DataFrame(missing_products)
    preview_df = pd.DataFrame(preview_rows)
    return created, skipped, updated, missing_products_df, preview_df


def build_month_check_lines_from_internet(df_internet: pd.DataFrame):
    """Готовит строки для ОДНОГО чека за месяц: группировка по артикулу.

    Возвращает:
      - lines_df: строки чека (article, name, product_id, qty, price, total)
      - missing_df: товары, которых нет в справочнике products (по article)
    """
    if df_internet is None or df_internet.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Ожидаем, что в df_internet есть как минимум: 'Артикул', 'Количество в заказе, шт.', 'Цена продажи, руб.'
    # Наименования колонок в отчёте WB могут немного отличаться — используем безопасный подбор.
    art_col = "Артикул" if "Артикул" in df_internet.columns else None
    name_col = "Наименование товара" if "Наименование товара" in df_internet.columns else None

    qty_col = None
    for cand in ["Количество в заказе, шт.", "Количество", "Кол-во", "qty"]:
        if cand in df_internet.columns:
            qty_col = cand
            break

    price_col = None
    for cand in ["Цена продажи, руб.", "Цена продажи", "sale_price", "price"]:
        if cand in df_internet.columns:
            price_col = cand
            break

    if art_col is None or qty_col is None or price_col is None:
        return pd.DataFrame(), pd.DataFrame({"Ошибка": ["Не найдены обязательные колонки (Артикул/Количество/Цена продажи)."]})

    tmp = df_internet.copy()
    tmp[art_col] = tmp[art_col].astype(str).str.strip()

    tmp["__qty"] = pd.to_numeric(tmp[qty_col], errors="coerce").fillna(0.0)
    tmp["__price"] = pd.to_numeric(tmp[price_col], errors="coerce").fillna(0.0)
    tmp["__total"] = (tmp["__qty"] * tmp["__price"]).round(2)

    # Защита от дублей: схлопываем одинаковый артикул в одну строку
    grp = tmp.groupby(art_col, as_index=False).agg(
        qty_sum=("__qty", "sum"),
        total_sum=("__total", "sum"),
    )

    grp["price_calc"] = grp.apply(lambda r: (r["total_sum"] / r["qty_sum"]) if r["qty_sum"] else 0.0, axis=1)
    grp["price_calc"] = grp["price_calc"].round(2)

    # Подтягиваем product_id и наименование из справочника
    products_df = pd.read_sql("SELECT id AS product_id, article, name FROM products", conn)
    products_df["article"] = products_df["article"].astype(str).str.strip()

    merged = grp.merge(products_df, left_on=art_col, right_on="article", how="left")

    missing = merged[merged["product_id"].isna()].copy()
    if not missing.empty:
        missing_df = missing[[art_col]].rename(columns={art_col: "Артикул"})
        if name_col and name_col in tmp.columns:
            # Попробуем добавить пример названия из исходника
            sample_names = tmp[[art_col, name_col]].dropna().drop_duplicates(subset=[art_col])
            sample_names = sample_names.rename(columns={art_col: "Артикул", name_col: "Наименование товара"})
            missing_df = missing_df.merge(sample_names, on="Артикул", how="left")
        missing_df["Причина"] = "Товар не найден в базе (products)"
    else:
        missing_df = pd.DataFrame()

    lines_df = merged[~merged["product_id"].isna()].copy()
    if lines_df.empty:
        return pd.DataFrame(), missing_df

    lines_df["product_id"] = lines_df["product_id"].astype(int)
    lines_df = lines_df.rename(columns={
        art_col: "Артикул",
        "name": "Наименование",
        "qty_sum": "Количество",
        "price_calc": "Цена",
        "total_sum": "Сумма"
    })

    lines_df = lines_df[["product_id", "Артикул", "Наименование", "Количество", "Цена", "Сумма"]]
    lines_df = lines_df.sort_values(["Артикул"])

    return lines_df, missing_df


def create_or_overwrite_month_sale(sale_number: str, sale_date_str: str, comment: str, overwrite: bool):
    """Создаёт один чек за месяц. Если overwrite=True и чек с таким номером уже есть — перезапишет позиции."""
    existing_id = get_sale_id_by_number(sale_number)
    if existing_id is None:
        c.execute(
            "INSERT INTO sales (sale_number, sale_date, comment) VALUES (?, ?, ?)",
            (sale_number, sale_date_str, comment),
        )
        conn.commit()
        return int(c.lastrowid)

    if not overwrite:
        return int(existing_id)

    # Перезаписываем: удаляем позиции и обновляем шапку
    c.execute("DELETE FROM sale_items WHERE sale_id=?", (existing_id,))
    c.execute(
        "UPDATE sales SET sale_date=?, comment=? WHERE id=?",
        (sale_date_str, comment, existing_id),
    )
    conn.commit()
    return int(existing_id)


def push_marketplace_month_to_db(lines_df: pd.DataFrame, sale_number: str, sale_date_obj, overwrite: bool = False):
    """Загружает ОДИН чек за месяц в sales/sale_items по подготовленным строкам."""
    if lines_df is None or lines_df.empty:
        return {"ok": False, "error": "Нет строк для загрузки."}

    sale_date_str = str(sale_date_obj)
    comment = "Интернет-торговля (месячный чек)"

    sale_id = create_or_overwrite_month_sale(sale_number, sale_date_str, comment, overwrite)

    # Если overwrite=False и чек уже был — не добавляем повторно (защита от дублей)
    if (not overwrite) and (get_sale_id_by_number(sale_number) == sale_id):
        # Проверим, есть ли уже позиции
        cnt = c.execute("SELECT COUNT(1) FROM sale_items WHERE sale_id=?", (sale_id,)).fetchone()[0]
        if cnt and cnt > 0:
            return {"ok": True, "sale_id": sale_id, "mode": "skipped_exists", "items_added": 0}

    items_added = 0
    for _, r in lines_df.iterrows():
        product_id = int(r["product_id"])
        qty = float(r["Количество"])
        price = float(r["Цена"])
        if qty == 0:
            continue
        gross_total = float(pd.to_numeric(r.get("Сумма", qty * price), errors="coerce") or 0)
        mp_fee = float(pd.to_numeric(r.get("Комиссия", 0), errors="coerce") or 0)
        mp_delivery = float(pd.to_numeric(r.get("Доставка", 0), errors="coerce") or 0)
        net_total = float(pd.to_numeric(r.get("Чистая_сумма", gross_total - mp_fee - mp_delivery), errors="coerce") or (gross_total - mp_fee - mp_delivery))
        add_sale_item_mp(sale_id, product_id, qty, price, gross_total, mp_fee, mp_delivery, net_total)
        items_added += 1

    return {"ok": True, "sale_id": sale_id, "mode": "created_or_updated", "items_added": items_added}

def build_internet_sales_from_files(file_wb, delivery_files):
    """
    Формирует таблицу интернет-продаж из:
      1) 'Файл детализации оказанных услуг' (продажи)
      2) одного или нескольких файлов 'Детализации услуг Европочты' (доставка)

    Возвращает:
      merged      — итоговая таблица
      stats       — словарь со статистикой матчей по доставке
      missing_df  — таблица заказов без найденной доставки
    """

    # -----------------------------
    # 1) ПРОДАЖИ (основной файл)
    # -----------------------------
    df1_raw = pd.read_excel(BytesIO(file_wb.getvalue() if hasattr(file_wb, 'getvalue') else file_wb), header=None)

    # В твоих образцах заголовок таблицы с продажами начинается примерно с 10-й строки.
    # Но на всякий случай попробуем найти строку, где есть "№ заказа" и "ШК".
    header_row_sales = None
    # Ищем строку заголовка по нормализованным названиям (устойчиво к переносам/пробелам/№->N)
    for i in range(min(len(df1_raw), 40)):
        row_norm = [_norm_col(x) for x in df1_raw.iloc[i].astype(str).tolist()]
        has_order = any(("заказ" in x and ("n" in x or "номер" in x or "id" in x)) for x in row_norm)
        has_shk = any(("шк" == x) or ("штрихкод" in x) or ("ean" in x) or ("barcode" in x) for x in row_norm)
        if has_order and has_shk:
            header_row_sales = i
            break
    if header_row_sales is None:
        header_row_sales = 9  # запасной вариант (как было)

    header1 = df1_raw.iloc[header_row_sales]
    data1 = df1_raw.iloc[header_row_sales + 1 :].copy()
    data1.columns = header1

    # Нормализуем названия колонок (убрать переносы/пробелы)
    data1.columns = [str(c).replace("\n", " ").replace("\r", " ").strip() for c in data1.columns]

    # Пробуем привести колонки к канону (устойчиво к разным вариантам заголовков)
    data1_std, missing_cols, _ = standardize_sales_columns(data1)
    if missing_cols:
        raise ValueError(
            "В файле продаж не найдены колонки: " + ", ".join(missing_cols)
            + "\n\nНайденные колонки: " + ", ".join([str(x) for x in data1.columns])
        )
    data1 = data1_std
    need_sales_cols = [
        "№ заказа",
        "ШК",
        "Наименование товара",
        "Количество в заказе, шт.",
        "Цена продажи, руб.",
        "К оплате с НДС, руб.",
        "Текущий статус",
    ]
    df_sales = data1[need_sales_cols].copy()

    # ФИЛЬТР: в продажи попадают ТОЛЬКО доставленные заказы
    df_sales["Текущий статус"] = (
        df_sales["Текущий статус"]
        .astype(str)
        .str.strip()
        .str.lower()
    )
    df_sales = df_sales[df_sales["Текущий статус"] == "доставлен"]

    df_sales = df_sales.rename(columns={
        "ШК": "Артикул",  # фактически это штрихкод (ШК)
        "К оплате с НДС, руб.": "Вознаграждение площадки",
        "Количество в заказе, шт.": "Количество"
    })

    # нормализация
    df_sales["_order_key"] = df_sales["№ заказа"].apply(order_match_key)
    df_sales = df_sales[df_sales["_order_key"] != ""]
    df_sales["Артикул"] = df_sales["Артикул"].astype(str).str.strip()
    df_sales["Количество"] = pd.to_numeric(df_sales["Количество"], errors="coerce").fillna(0.0)
    df_sales["Цена продажи, руб."] = pd.to_numeric(df_sales["Цена продажи, руб."], errors="coerce").fillna(0.0)
    df_sales["Вознаграждение площадки"] = pd.to_numeric(df_sales["Вознаграждение площадки"], errors="coerce").fillna(0.0)

    # -----------------------------
    # 2) ДОСТАВКА (Европочта, несколько файлов)
    # -----------------------------
    delivery_frames = []

    if delivery_files is None:
        delivery_files = []
    if not isinstance(delivery_files, (list, tuple)):
        delivery_files = [delivery_files]

    for f in delivery_files:
        if f is None:
            continue

        df2_raw = pd.read_excel(BytesIO(f.getvalue() if hasattr(f, 'getvalue') else f), header=None)

        # ищем строку заголовка (где есть "№ отправления" или "№ заказа")
        header_row_del = None
        for i in range(min(len(df2_raw), 30)):
            row = df2_raw.iloc[i].astype(str).tolist()
            if ("№ отправления" in row) or ("№ заказа" in row) or ("Номер заказа" in row):
                header_row_del = i
                break
        if header_row_del is None:
            header_row_del = 8  # как было

        header2 = df2_raw.iloc[header_row_del]
        data2 = df2_raw.iloc[header_row_del + 1 :].copy()
        data2.columns = header2

        # колонка идентификатора заказа
        order_col = None
        for cand in ["№ отправления", "№ заказа", "Номер заказа"]:
            if cand in data2.columns:
                order_col = cand
                break
        if order_col is None:
            # не удалось распознать файл — пропускаем
            continue

        # колонка стоимости
        cost_col = None
        for cand in ["Сумма с НДС, руб.", "Стоимость доставки", "Сумма, руб.", "Стоимость, руб."]:
            if cand in data2.columns:
                cost_col = cand
                break
        if cost_col is None:
            continue

        df_delivery = data2[[order_col, cost_col]].copy()
        df_delivery = df_delivery.rename(columns={
            order_col: "№ заказа",
            cost_col: "Стоимость доставки"
        })

        df_delivery["_order_key"] = df_delivery["№ заказа"].apply(order_match_key)
        df_delivery["Стоимость доставки"] = pd.to_numeric(df_delivery["Стоимость доставки"], errors="coerce").fillna(0.0)

        # отфильтруем мусорные строки
        df_delivery = df_delivery[(df_delivery["_order_key"] != "NAN") & (df_delivery["_order_key"] != "")]

        delivery_frames.append(df_delivery)

    if delivery_frames:
        df_delivery_all = pd.concat(delivery_frames, ignore_index=True)
        # если один заказ встречается в разных файлах — суммируем
        df_delivery_agg = (
            df_delivery_all
            .groupby("_order_key", as_index=False)["Стоимость доставки"]
            .sum()
        )
    else:
        df_delivery_agg = pd.DataFrame(columns=["№ заказа", "Стоимость доставки"])

    # -----------------------------
    # 3) ОБЪЕДИНЕНИЕ + СТАТИСТИКА
    # -----------------------------
    merged = df_sales.merge(df_delivery_agg, on="_order_key", how="left")

    # для удобства: NaN -> 0, но до этого посчитаем матчи
    
    total_orders = merged["_order_key"].dropna().astype(str).loc[lambda s: (s != "") & (s != "NAN")].nunique()
    matched_orders = (
    merged.loc[merged["Стоимость доставки"].notna(), "_order_key"]
    .dropna()
    .astype(str)
    .loc[lambda s: (s != "") & (s != "NAN")]
    .nunique()
    )
    missing_orders = (
    merged.loc[merged["Стоимость доставки"].isna(), "_order_key"]
    .dropna()
    .astype(str)
    .loc[lambda s: (s != "") & (s != "NAN")]
    .drop_duplicates()
    .sort_values()
    )

    merged["Стоимость доставки"] = merged["Стоимость доставки"].fillna(0.0)

    # Распределяем доставку по строкам заказа пропорционально сумме строки
    merged["gross_total"] = (merged["Количество"] * merged["Цена продажи, руб."]).round(2)
    order_sum = merged.groupby("_order_key")["gross_total"].transform("sum").replace(0, 1.0)
    share = merged["gross_total"] / order_sum

    delivery_per_order = merged.groupby("_order_key")["Стоимость доставки"].transform("max").fillna(0.0)
    merged["Доставка"] = (delivery_per_order * share).round(2)

    # Комиссия (вознаграждение площадки): если одинаковая на весь заказ — распределяем, иначе считаем построчно
    fee_raw = merged["Вознаграждение площадки"].fillna(0.0)
    rows_in_order = merged.groupby("_order_key")["Артикул"].transform("size")
    fee_nuniq = merged.groupby("_order_key")["Вознаграждение площадки"].transform("nunique").fillna(0)
    fee_per_order = merged.groupby("_order_key")["Вознаграждение площадки"].transform("max").fillna(0.0)
    is_order_level_fee = (rows_in_order > 1) & (fee_nuniq == 1)

    merged["Комиссия"] = pd.to_numeric(
    (fee_per_order * share).where(is_order_level_fee, fee_raw),
    errors="coerce"
    ).fillna(0.0).round(2)

    merged["Сумма"] = merged["gross_total"]
    merged["Чистая_сумма"] = (merged["Сумма"] - merged["Комиссия"] - merged["Доставка"]).round(2)

    stats = {
    "total_orders": int(total_orders),
    "matched_orders": int(matched_orders),
    "missing_orders": int(len(missing_orders)),
    "delivery_total_files": float(pd.to_numeric(df_delivery_agg["Стоимость доставки"], errors="coerce").fillna(0).sum()) if not df_delivery_agg.empty else 0.0,
    "delivery_total_applied": float(pd.to_numeric(delivery_per_order, errors="coerce").fillna(0).drop_duplicates().sum()) if total_orders else 0.0,
    }

    missing_df = pd.DataFrame({"№ заказа (ключ)": missing_orders})
    return merged, stats, missing_df


    # -----------------------------
    # 1. ПРИХОДЫ (партии) ЗА ГОД
    # -----------------------------
    df_items = pd.read_sql("""
        SELECT
            ii.id AS item_id,
            ii.product_id,
            p.name AS product_name,
            ii.qty AS qty_in,
            ii.price AS price_in,
            i.invoice_number,
            i.invoice_date,
            i.supplier
        FROM invoice_items ii
        JOIN invoices i ON i.id = ii.invoice_id
        JOIN products p ON p.id = ii.product_id
    """, conn)

    if df_items.empty:
        return None

    df_items["invoice_date_dt"] = pd.to_datetime(df_items["invoice_date"])
    df_items["year"] = df_items["invoice_date_dt"].dt.year
    df_items["month"] = df_items["invoice_date_dt"].dt.month

    # Оставляем только выбранный год
    df_items = df_items[df_items["year"] == selected_year].copy()
    if df_items.empty:
        return None
    


    # -----------------------------
    # 2. ПРОДАЖИ ЗА IV КВАРТАЛ
    #    (отдельно, по году и месяцу 10–12)
    # -----------------------------
    df_sales = pd.read_sql("""
    SELECT 
        s.id AS sale_id,
        s.sale_number,
        s.sale_date,
        p.id AS product_id,
        p.article,
        p.name,
        p.unit,
        si.qty,
        si.price,
        si.total
    FROM sale_items si
    JOIN sales s ON s.id = si.sale_id
    JOIN products p ON p.id = si.product_id
    WHERE s.sale_date BETWEEN ? AND ?
    ORDER BY s.sale_date, s.sale_number
    """, conn)

    if not df_sales.empty:
        df_sales["sale_date_dt"] = pd.to_datetime(df_sales["sale_date"])
        df_sales["sale_year"] = df_sales["sale_date_dt"].dt.year
        df_sales["sale_month"] = df_sales["sale_date_dt"].dt.month

        # Берём только продажи выбранного года и IV квартала (10–12 месяц)
        df_sales = df_sales[
            (df_sales["sale_year"] == selected_year) &
            (df_sales["sale_month"].between(10, 12))
        ].copy()
    else:
        df_sales = pd.DataFrame(columns=[
            "product_id", "qty_sold", "sale_price",
            "sale_number", "sale_date", "sale_date_dt",
            "sale_year", "sale_month"
        ])

    # -----------------------------
    # 3. ГОТОВИМ FIFO ПО КАЖДОМУ ТОВАРУ
    # -----------------------------
    # Структуры:
    #  - product_batches[product_id] = [item_id1, item_id2, ...] (в порядке даты прихода)
    #  - batch_info[item_id] = данные партии + результаты списаний/остатков
    product_batches = defaultdict(list)
    batch_info = {}

    # Сначала заполним партии (приходы)
    df_items_sorted = df_items.sort_values(
        by=["product_id", "invoice_date_dt", "item_id"]
    )

    for _, r in df_items_sorted.iterrows():
        item_id = int(r["item_id"])
        pid = int(r["product_id"])
        qty_in = float(r["qty_in"])
        price_in = float(r["price_in"])

        product_batches[pid].append(item_id)
        batch_info[item_id] = {
            "product_id": pid,
            "product_name": r["product_name"],   # ← ДОБАВЛЕНО
            "qty_in": qty_in,
            "price_in": price_in,
            "invoice_number": r["invoice_number"],
            "invoice_date": r["invoice_date"],
            "supplier": r["supplier"] or "",
            "month": int(r["month"]),
            "qty_remaining": qty_in,
            "sales_alloc": OrderedDict(),
            "expense": 0.0,
        }

    # Теперь распределяем продажи по партиям (FIFO) для IV квартала
    if not df_sales.empty:
        df_sales_sorted = df_sales.sort_values(
            by=["product_id", "sale_date_dt"]
        )

        for _, srow in df_sales_sorted.iterrows():
            pid = int(srow["product_id"])
            sale_qty = float(srow["qty_sold"])
            sale_price = float(srow["sale_price"])
            sale_number = str(srow["sale_number"])
            sale_date = srow["sale_date_dt"]

            if pid not in product_batches:
                # Продажа товара, по которому нет приходов в выбранном году – пропускаем
                continue

            # FIFO по партиям этого товара
            for item_id in product_batches[pid]:
                if sale_qty <= 0:
                    break

                binfo = batch_info[item_id]
                if binfo["qty_remaining"] <= 0:
                    continue

                # Сколько можно списать из этой партии
                alloc = min(sale_qty, binfo["qty_remaining"])
                if alloc <= 0:
                    continue

                sale_qty -= alloc
                binfo["qty_remaining"] -= alloc

                # Расходы: списываем по закупочной цене этой партии
                binfo["expense"] += alloc * binfo["price_in"]

                # Фиксируем продажу за этой партией
                if sale_number not in binfo["sales_alloc"]:
                    binfo["sales_alloc"][sale_number] = {
                        "qty": 0.0,
                        "price": sale_price,
                        "first_date": sale_date,
                    }
                binfo["sales_alloc"][sale_number]["qty"] += alloc

            # Если после всех партий sale_qty > 0 – значит продали больше, чем пришло в этом году.
            # По хорошему, такого быть не должно, поэтому тут ничего не делаем, просто игнор.

    # -----------------------------
    # 4. ЗАПИСЫВАЕМ ВСЁ В EXCEL
    # -----------------------------
    # Находим первую свободную строку, начиная с 12-й
    row = 12
    while ws[f"A{row}"].value not in (None, ""):
        row += 1

    # Будем проходить по партиям (df_items_sorted) — строго в том же порядке
    for _, r in df_items_sorted.iterrows():
        item_id = int(r["item_id"])
        binfo = batch_info[item_id]

        supplier = binfo["supplier"]
        inv_num = binfo["invoice_number"]
        inv_date = binfo["invoice_date"]
        price_in = binfo["price_in"]
        qty_in = binfo["qty_in"]
        month = binfo["month"]

        # ---- A, B ----
        ws[f"A{row}"] = f"{supplier}, №{inv_num} от {inv_date}"
        # Наименование товара мы в данный момент в df не брали, но можно легко добавить при желании.
        # Пока пишем product_id как заглушку или оставляем пустым.
        # Лучше пока оставить пустым, пока ты не скажешь, что именно нужно.
        ws[f"B{row}"] = binfo["product_name"]

        # ---- C, D ----
        # Остаток на начало года — для первого года = 0
        ws[f"C{row}"] = 0
        ws[f"D{row}"] = 0

        # ---- IV квартал: приход (AU, AV, AW) ----
        # Только если эта партия пришла в IV квартале (10–12 месяцы)
        if month in (10, 11, 12):
            ws[f"AU{row}"] = float(price_in)            # Цена закупки
            ws[f"AV{row}"] = float(qty_in)              # Количество закупки
            ws[f"AW{row}"] = float(price_in * qty_in)   # Стоимость закупки

            # ---- ПРОДАЖИ по данной партии (BC–BF) ----
            sales_alloc = binfo["sales_alloc"]

            if sales_alloc:
                # Сортируем продажи по дате первой продажи
                alloc_list = sorted(
                    sales_alloc.items(),
                    key=lambda kv: kv[1]["first_date"]
                )

                # BC: номера чеков
                sale_numbers = [sn for sn, _ in alloc_list]
                ws[f"BC{row}"] = ", ".join(sale_numbers)

                # BE: количества по каждому чеку
                sale_qtys = [str(round(info["qty"], 2)) for _, info in alloc_list]
                total_sold = sum(info["qty"] for _, info in alloc_list)
                ws[f"BE{row}"] = float(total_sold)

                # BD: цены продажи — одно значение, если все одинаковые; иначе — через запятую
                prices = [round(info["price"], 2) for _, info in alloc_list]
                unique_prices = []
                for p in prices:
                    if p not in unique_prices:
                        unique_prices.append(p)

                if len(unique_prices) == 1:
                    ws[f"BD{row}"] = float(unique_prices[0])
                else:
                    ws[f"BD{row}"] = ", ".join(str(p) for p in unique_prices)

                # BF: расходы по закупочной цене этой партии
                ws[f"BF{row}"] = float(binfo["expense"])
            else:
                # Не было продаж по этой партии
                ws[f"BC{row}"] = ""
                ws[f"BD{row}"] = ""
                ws[f"BE{row}"] = ""
                ws[f"BF{row}"] = 0

            # ---- Остаток по партии (BG, BH) ----
            qty_rem = binfo["qty_remaining"]
            ws[f"BG{row}"] = float(qty_rem)
            ws[f"BH{row}"] = float(qty_rem * price_in)
        # Если партия не в IV квартале — AU–BH пока не трогаем (оставляем пустым)

        row += 1

    # -----------------------------
    # 5. Сохраняем в буфер
    # -----------------------------
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ===================================
# STREAMLIT ИНТЕРФЕЙС
# ===================================
st.set_page_config(page_title="Складской учёт", page_icon="📦", layout="wide")


# --- session state for marketplace monthly draft ---
from datetime import date as _date
if "mp_draft_ready" not in st.session_state:
    st.session_state.mp_draft_ready = False
if "mp_draft_lines_df" not in st.session_state:
    st.session_state.mp_draft_lines_df = None
if "mp_draft_missing_products_df" not in st.session_state:
    st.session_state.mp_draft_missing_products_df = None
if "mp_draft_stats" not in st.session_state:
    st.session_state.mp_draft_stats = {}
if "mp_draft_missing_delivery_df" not in st.session_state:
    st.session_state.mp_draft_missing_delivery_df = None
if "mp_sale_number" not in st.session_state:
    st.session_state.mp_sale_number = ""
if "mp_sale_date" not in st.session_state:
    st.session_state.mp_sale_date = _date.today()
if "mp_overwrite" not in st.session_state:
    st.session_state.mp_overwrite = False

st.title("📦 Складской учёт")

menu = st.sidebar.radio("Навигация", [
    "➕ Добавить накладную",
    "📚 Просмотр базы",
    "🏷️ Склад",
    "🛒 Продажи",
    "📊 Отчёты"
])


# ---------------------------------------------------
# ➕ 1. Добавление накладной
# ---------------------------------------------------
if menu == "➕ Добавить накладную":
    st.subheader("Добавление новой накладной")

    col1, col2, col3 = st.columns(3)
    with col1:
        inv_number = st.text_input("Номер накладной *")
    with col2:
        inv_date = st.date_input("Дата накладной", value=date.today())
    with col3:
        supplier = st.text_input("Поставщик")

    if st.button("Создать накладную"):
        if not inv_number:
            st.error("Введите номер накладной.")
        else:
            inv_id = insert_invoice(inv_number, str(inv_date), supplier)
            st.success(f"Накладная №{inv_number} от {inv_date} создана (ID={inv_id})")

    st.markdown("---")
    st.subheader("Добавление товаров в накладную")

    invoices_df = get_invoices_df()
    if invoices_df.empty:
        st.info("Нет накладных. Сначала создайте накладную.")
    else:
        inv_label = invoices_df.apply(lambda x: f"{x['invoice_number']} — {x['invoice_date']}", axis=1)
        selected_inv = st.selectbox("Выберите накладную", inv_label)
        invoice_id = int(invoices_df.loc[inv_label == selected_inv, "id"].iloc[0])

        search = st.text_input("🔍 Поиск товара (артикул или название)")
        products_df = get_products(search)

        if products_df.empty:
            st.warning("Не найдено товаров. Добавьте новый.")
            with st.expander("➕ Добавить новый товар"):
                art = st.text_input("Артикул *")
                name = st.text_input("Наименование *")
                unit = st.text_input("Единица измерения", "шт")
                if st.button("Добавить товар"):
                    if not art or not name:
                        st.error("Введите артикул и наименование.")
                    else:
                        insert_product(art, name, unit)
                        st.success("Товар добавлен.")
                        st.rerun()
        else:
            label = products_df.apply(lambda r: f"{r['name']} ({r['article']})", axis=1)
            sel = st.selectbox("Выберите товар", label)
            product_id = int(products_df.loc[label == sel, "id"].iloc[0])

            col1, col2, col3 = st.columns(3)
            with col1:
                qty = st.number_input("Количество", min_value=0.0, step=1.0)
            with col2:
                price = st.number_input("Цена", min_value=0.0, step=0.01)
            with col3:
                vat = st.number_input("НДС (%)", min_value=0.0, step=1.0, value=20.0)

            if st.button("Добавить в накладную"):
                add_item(invoice_id, product_id, qty, price, vat)
                st.success("Строка добавлена.")
                st.rerun()

        items_df = get_invoice_items_df(invoice_id)
        if not items_df.empty:
            st.dataframe(items_df, use_container_width=True)
            q, s, sv = invoice_totals(invoice_id)
            st.markdown(f"**Итого:** Кол-во: {q:.2f}, Сумма без НДС: {s:.2f}, С НДС: {sv:.2f}")

# ---------------------------------------------------
# 📚 3. Просмотр базы
# ---------------------------------------------------
if menu == "📚 Просмотр базы":
    st.subheader("Просмотр базы данных")
    view = st.radio("Выберите раздел", [
        "Товары",
        "Накладные",
        "Позиции накладных",
        "Продажи (чеки)"
    ])

    # ---- ТОВАРЫ ----
    if view == "Товары":
        search = st.text_input("Поиск товара (название / артикул)")
        df = get_products(search)
        st.dataframe(df, use_container_width=True)

        if not df.empty:
            st.markdown("### ✏️ Редактирование товара")
            names = df["name"].tolist()
            selected_name = st.selectbox("Выберите товар", names)
            product = df[df["name"] == selected_name].iloc[0]

            new_article = st.text_input("Артикул", value=product["article"])
            new_name = st.text_input("Наименование", value=product["name"])
            new_unit = st.text_input("Единица изм.", value=product["unit"] or "")

            if st.button("💾 Сохранить изменения"):
                c.execute(
                    "UPDATE products SET article=?, name=?, unit=? WHERE id=?",
                    (new_article, new_name, new_unit, int(product["id"]))
                )
                conn.commit()
                st.success("Изменения сохранены.")
                st.rerun()

     # ---- НАКЛАДНЫЕ ----
    elif view == "Накладные":
        df = get_invoices_df()
        if df.empty:
            st.info("Нет накладных.")
        else:
            # Добавляем суммы к каждой накладной
            df["Сумма_накладной"] = df["id"].apply(lambda x: get_invoice_sum(x))

            # Фильтры
            st.markdown("### 🔍 Фильтры")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                dmin = pd.to_datetime(df["invoice_date"]).min().date()
                dmax = pd.to_datetime(df["invoice_date"]).max().date()
                date_from = st.date_input("Дата с", value=dmin)
            with col2:
                date_to = st.date_input("Дата по", value=dmax)
            with col3:
                num_filter = st.text_input("Номер накладной содержит")
            with col4:
                supplier_filter = st.text_input("Поставщик содержит")

            col5, col6 = st.columns(2)
            with col5:
                min_sum = st.number_input("Мин. сумма", min_value=0.0, step=10.0)
            with col6:
                max_sum = st.number_input("Макс. сумма", min_value=0.0, step=10.0, value=float(df["Сумма_накладной"].max() or 0))

            df = df[
                (pd.to_datetime(df["invoice_date"]).dt.date >= date_from) &
                (pd.to_datetime(df["invoice_date"]).dt.date <= date_to) &
                (df["Сумма_накладной"].between(min_sum, max_sum))
            ]
            if num_filter:
                df = df[df["invoice_number"].str.contains(num_filter, case=False, na=False)]
            if supplier_filter:
                df = df[df["supplier"].astype(str).str.contains(supplier_filter, case=False, na=False)]

            st.dataframe(df, use_container_width=True)

            if not df.empty:
                selected_label = df.apply(lambda r: f"{r['invoice_number']} — {r['invoice_date']}", axis=1)
                selected = st.selectbox("Выберите накладную для редактирования", selected_label)
                invoice_id = int(df.loc[selected_label == selected, "id"].iloc[0])
                items_df = get_invoice_items_df(invoice_id)
                if items_df.empty:
                    st.info("Нет позиций в накладной.")
                else:
                    for _, row in items_df.iterrows():
                        with st.expander(f"{row['Наименование']} ({row['Артикул']})"):
                            c1, c2, c3 = st.columns(3)
                            with c1:
                                new_qty = st.number_input("Количество", value=float(row["Количество"]), step=1.0, key=f"q_{row['id']}")
                            with c2:
                                new_price = st.number_input("Цена", value=float(row["Цена"]), step=0.01, key=f"p_{row['id']}")
                            with c3:
                                new_vat = st.number_input("НДС (%)", value=float(row["НДС"]), step=1.0, key=f"v_{row['id']}")
                            if st.button("💾 Сохранить", key=f"save_{row['id']}"):
                                update_invoice_item(int(row["id"]), new_qty, new_price, new_vat)
                                st.success("Изменено.")
                                st.rerun()
                            if st.button("🗑️ Удалить", key=f"del_{row['id']}"):
                                delete_invoice_item(int(row["id"]))
                                st.warning("Удалено.")
                                st.rerun()
                    q, s, sv = invoice_totals(invoice_id)
                    st.markdown(f"**Итого:** Кол-во: {q:.2f}, Сумма без НДС: {s:.2f}, С НДС: {sv:.2f}")

            # --- Редактирование данных накладной ---
            st.markdown("### ✏️ Редактирование параметров накладной")
            inv_row = df[df["id"] == invoice_id].iloc[0]
            col1, col2, col3 = st.columns(3)

            with col1:
                new_number = st.text_input("Номер накладной", value=inv_row["invoice_number"])
            with col2:
                new_date = st.date_input("Дата накладной", value=pd.to_datetime(inv_row["invoice_date"]).date())
            with col3:
                new_supplier = st.text_input("Поставщик", value=inv_row["supplier"] or "")

            if st.button("💾 Сохранить изменения по накладной"):
                c.execute("""
                    UPDATE invoices
                    SET invoice_number = ?, invoice_date = ?, supplier = ?
                    WHERE id = ?
                """, (new_number, str(new_date), new_supplier, invoice_id))
                conn.commit()
                st.success("Изменения сохранены.")
                st.rerun()

    # ---- ПОЗИЦИИ НАКЛАДНЫХ ----
    elif view == "Позиции накладных":
        df = pd.read_sql("""
            SELECT i.invoice_number AS Накладная, i.invoice_date AS Дата, i.supplier AS Поставщик,
                   p.article AS Артикул, p.name AS Наименование, ii.qty AS Количество,
                   ii.price AS Цена, ii.total AS Сумма, ii.total_with_vat AS Сумма_с_НДС
            FROM invoice_items ii
            JOIN invoices i ON i.id = ii.invoice_id
            JOIN products p ON p.id = ii.product_id
            ORDER BY i.invoice_date DESC
        """, conn)
        st.dataframe(df, use_container_width=True)

    

    # ---- ПРОДАЖИ (ЧЕКИ) ----
    elif view == "Продажи (чеки)":
        sales_df = get_sales_df()
        if sales_df.empty:
            st.info("Пока нет чеков.")
        else:
            st.dataframe(sales_df, use_container_width=True)

            st.markdown("### ✏️ Редактирование чека")
            sale_label = sales_df.apply(lambda r: f"{r['sale_number']} — {r['sale_date']}", axis=1)
            selected_sale = st.selectbox("Выберите чек для редактирования", sale_label)

            # Находим выбранный чек по id
            sale_id = int(sales_df.loc[sale_label == selected_sale, "id"].iloc[0])
            sale_row = sales_df[sales_df["id"] == sale_id].iloc[0]

            col1, col2, col3 = st.columns(3)
            with col1:
                new_number = st.text_input("Номер чека", value=sale_row["sale_number"])
            with col2:
                new_date = st.date_input(
                    "Дата продажи",
                    value=pd.to_datetime(sale_row["sale_date"]).date()
                )
            with col3:
                new_comment = st.text_input("Комментарий", value=sale_row["comment"] or "")

            c1, c2 = st.columns(2)
            with c1:
                if st.button("💾 Сохранить изменения по чеку"):
                    c.execute(
                        "UPDATE sales SET sale_number=?, sale_date=?, comment=? WHERE id=?",
                        (new_number, str(new_date), new_comment, sale_id)
                    )
                    conn.commit()
                    try:
                        recalc_fifo_for_sale(conn, sale_id)
                    except Exception:
                        pass
                    st.success("Изменения по чеку сохранены.")
                    st.rerun()

            with c2:
                if st.button("🔁 Пересчитать FIFO для этого чека"):
                    try:
                        logger.info("Manual FIFO recalculation requested for sale_id=%s", sale_id)
                        recalc_fifo_for_sale(conn, sale_id)
                        logger.info("Manual FIFO recalculation finished for sale_id=%s", sale_id)
                        st.success("FIFO пересчитан для этого чека.")
                    except Exception as e:
                        logger.exception("Error during manual FIFO recalc for sale_id=%s: %s", sale_id, e)
                        st.error(f"Ошибка при пересчёте FIFO: {e}")
                    st.rerun()

            with c2:
                if st.button("🗑️ Удалить чек целиком"):
                    delete_sale(sale_id)
                    st.warning("Чек и все его позиции удалены.")
                    st.rerun()

            st.markdown("### 📋 Позиции выбранного чека")
            try:
                show_fifo_for_sale(conn, sale_id)
            except Exception:
                pass
            sale_items = get_sale_items_df(sale_id)
            if sale_items.empty:
                st.info("В этом чеке нет товаров.")
            else:
                for _, row in sale_items.iterrows():
                    with st.expander(f"{row['Наименование']} ({row['Артикул']})"):
                        col1, col2 = st.columns(2)
                        with col1:
                            new_qty = st.number_input(
                                "Количество",
                                value=float(row["Количество"]),
                                step=1.0,
                                key=f"sq_{row['id']}"
                            )
                        with col2:
                            new_price = st.number_input(
                                "Цена продажи",
                                value=float(row["Цена"]),
                                step=0.01,
                                key=f"sp_{row['id']}"
                            )

                        new_total = new_qty * new_price
                        st.write(f"**Сумма:** {new_total:.2f}")

                        cc1, cc2 = st.columns(2)
                        with cc1:
                            if st.button("💾 Сохранить позицию", key=f"ssave_{row['id']}"):
                                update_sale_item(int(row["id"]), new_qty, new_price)
                                st.success("Позиция сохранена.")
                                st.rerun()
                        with cc2:
                            if st.button("🗑️ Удалить позицию", key=f"sdel_{row['id']}"):
                                delete_sale_item(int(row["id"]))
                                st.warning("Позиция удалена.")
                                st.rerun()

                q, s = sale_totals(sale_id)
                st.markdown(f"**Итого по чеку:** Кол-во: {q:.2f}, Сумма: {s:.2f}")

# ---------------------------------------------------
# 🏷️ 4. Склад — остатки
# ---------------------------------------------------
elif menu == "🏷️ Склад":
    st.subheader("📦 Остатки товаров на дату")

    # --- состояние выбора (для двусторонней навигации) ---
    if "stock_selected_pid" not in st.session_state:
        st.session_state.stock_selected_pid = None
    if "stock_selected_sale_id" not in st.session_state:
        st.session_state.stock_selected_sale_id = None

    selected_date = st.date_input("Показать остатки на дату", value=date.today())
    search_text = st.text_input("Поиск (название / артикул)")

    df_stock = pd.read_sql("""
        SELECT 
            p.id AS product_id,
            p.article AS Артикул,
            p.name AS Наименование,
            p.unit AS Ед,
            SUM(ii.qty) AS Приход
        FROM invoice_items ii
        JOIN products p ON p.id = ii.product_id
        JOIN invoices i ON i.id = ii.invoice_id
        WHERE DATE(i.invoice_date) <= DATE(?)
        GROUP BY p.id, p.article, p.name, p.unit
        ORDER BY p.name
    """, conn, params=(str(selected_date),))

    # вычтем расход
    df_out = pd.read_sql("""
        SELECT p.id AS product_id, SUM(si.qty) AS Расход
        FROM sale_items si
        JOIN sales s ON s.id = si.sale_id
        JOIN products p ON p.id = si.product_id
        WHERE DATE(s.sale_date) <= DATE(?)
        GROUP BY p.id
    """, conn, params=(str(selected_date),))

    df_stock = pd.merge(df_stock, df_out, on="product_id", how="left").fillna(0)
    df_stock["Остаток"] = df_stock["Приход"] - df_stock["Расход"]

    if search_text:
        df_stock = df_stock[
            df_stock["Наименование"].str.contains(search_text, case=False, na=False) |
            df_stock["Артикул"].astype(str).str.contains(search_text, case=False, na=False)
        ]

    if df_stock.empty:
        st.info("Нет данных по остаткам на выбранную дату.")
    else:
        top_left, top_right = st.columns([3, 1])
        with top_right:
            if st.button("🧹 Сбросить выбор"):
                st.session_state.stock_selected_pid = None
                st.session_state.stock_selected_sale_id = None
                st.rerun()

        # Таблица остатков с возможностью выбрать товар (drill-down)
        event = st.dataframe(
            df_stock[["Артикул", "Наименование", "Ед", "Приход", "Расход", "Остаток"]],
            use_container_width=True,
            hide_index=True,
            on_select="rerun",
            selection_mode="single-row",
        )

        total = df_stock["Остаток"].sum()
        st.markdown(f"**Общий остаток (в штуках): {total:,.2f}**")

        # 1) выбор из клика
        selected_pid = None
        try:
            if event and event.selection and event.selection.get("rows"):
                ridx = event.selection["rows"][0]
                selected_pid = int(df_stock.iloc[ridx]["product_id"])
                st.session_state.stock_selected_pid = selected_pid
                # при выборе товара сбрасываем выбранный чек
                st.session_state.stock_selected_sale_id = None
        except Exception:
            selected_pid = None

        # 2) или из состояния (когда пришли "обратно" из чека)
        if selected_pid is None and st.session_state.stock_selected_pid is not None:
            selected_pid = int(st.session_state.stock_selected_pid)

        # --- Детализация по выбранному товару ---
        if selected_pid is not None:
            st.markdown("---")

            tabs = st.tabs(["🧾 Продажи (чеки)", "📥 Приходы (накладные)", "📈 Движение по датам"])

            # -----------------------------
            # 🧾 Продажи (чеки)
            # -----------------------------
            with tabs[0]:
                st.subheader("🧾 Продажи выбранного товара (чеки) до выбранной даты")

                df_sales_drill = pd.read_sql("""
                    SELECT
                        s.id          AS sale_id,
                        s.sale_date   AS Дата,
                        s.sale_number AS Чек,
                        COALESCE(s.comment, '') AS Комментарий,
                        SUM(si.qty)   AS Количество,
                        ROUND(AVG(si.price), 2) AS Цена_средняя,
                        ROUND(SUM(si.total), 2) AS Сумма
                    FROM sale_items si
                    JOIN sales s ON s.id = si.sale_id
                    WHERE si.product_id = ?
                      AND DATE(s.sale_date) <= DATE(?)
                    GROUP BY s.id, s.sale_date, s.sale_number, s.comment
                    ORDER BY DATE(s.sale_date) DESC, s.sale_number DESC
                """, conn, params=(selected_pid, str(selected_date)))

                if df_sales_drill.empty:
                    st.info("Продаж этого товара до выбранной даты нет.")
                else:
                    # таблица чеков с выбором
                    df_sales_view = df_sales_drill.drop(columns=["sale_id"]).copy()
                    ev_sale = st.dataframe(
                        df_sales_view,
                        use_container_width=True,
                        hide_index=True,
                        on_select="rerun",
                        selection_mode="single-row",
                    )

                    colA, colB = st.columns(2)
                    with colA:
                        st.metric("Итого продано (шт.)", f"{df_sales_drill['Количество'].sum():.2f}")
                    with colB:
                        st.metric("Итого сумма", f"{df_sales_drill['Сумма'].sum():.2f}")

                    # выбор чека
                    sale_id = None
                    try:
                        if ev_sale and ev_sale.selection and ev_sale.selection.get("rows"):
                            sidx = ev_sale.selection["rows"][0]
                            sale_id = int(df_sales_drill.iloc[sidx]["sale_id"])
                            st.session_state.stock_selected_sale_id = sale_id
                    except Exception:
                        sale_id = None

                    if sale_id is None and st.session_state.stock_selected_sale_id is not None:
                        sale_id = int(st.session_state.stock_selected_sale_id)

                    # --- двусторонняя навигация: из чека обратно в товар ---
                    if sale_id is not None:
                        st.markdown("#### 🧺 Состав выбранного чека (клик по товару — вернуться к нему)")
                        df_sale_items_full = pd.read_sql("""
                            SELECT
                                p.id AS product_id,
                                p.article AS Артикул,
                                p.name AS Наименование,
                                p.unit AS Ед,
                                si.qty AS Количество,
                                si.price AS Цена,
                                si.total AS Сумма
                            FROM sale_items si
                            JOIN products p ON p.id = si.product_id
                            WHERE si.sale_id = ?
                            ORDER BY si.id
                        """, conn, params=(sale_id,))

                        if df_sale_items_full.empty:
                            st.info("В этом чеке нет позиций.")
                        else:
                            df_sale_items_view = df_sale_items_full.drop(columns=["product_id"]).copy()
                            ev_item = st.dataframe(
                                df_sale_items_view,
                                use_container_width=True,
                                hide_index=True,
                                on_select="rerun",
                                selection_mode="single-row",
                            )

                            try:
                                if ev_item and ev_item.selection and ev_item.selection.get("rows"):
                                    iidx = ev_item.selection["rows"][0]
                                    pid2 = int(df_sale_items_full.iloc[iidx]["product_id"])
                                    st.session_state.stock_selected_pid = pid2
                                    st.session_state.stock_selected_sale_id = sale_id
                                    st.rerun()
                            except Exception:
                                pass

            # -----------------------------
            # 📥 Приходы (накладные)
            # -----------------------------
            with tabs[1]:
                st.subheader("📥 Приходы выбранного товара (накладные) до выбранной даты")

                df_in_drill = pd.read_sql("""
                    SELECT
                        i.invoice_date   AS Дата,
                        i.invoice_number AS Накладная,
                        COALESCE(i.supplier, '') AS Поставщик,
                        SUM(ii.qty)      AS Количество,
                        ROUND(AVG(ii.price), 2) AS Цена_средняя,
                        ROUND(SUM(ii.total_with_vat), 2) AS Сумма_с_НДС
                    FROM invoice_items ii
                    JOIN invoices i ON i.id = ii.invoice_id
                    WHERE ii.product_id = ?
                      AND DATE(i.invoice_date) <= DATE(?)
                    GROUP BY i.invoice_date, i.invoice_number, i.supplier
                    ORDER BY DATE(i.invoice_date) DESC, i.invoice_number DESC
                """, conn, params=(selected_pid, str(selected_date)))

                if df_in_drill.empty:
                    st.info("Приходов этого товара до выбранной даты нет.")
                else:
                    st.dataframe(df_in_drill, use_container_width=True, hide_index=True)
                    colI1, colI2 = st.columns(2)
                    with colI1:
                        st.metric("Итого приход (шт.)", f"{df_in_drill['Количество'].sum():.2f}")
                    with colI2:
                        st.metric("Итого сумма с НДС", f"{df_in_drill['Сумма_с_НДС'].sum():.2f}")

            # -----------------------------
            # 📈 Движение по датам
            # -----------------------------
            with tabs[2]:
                st.subheader("📈 Движение по датам (приход − расход)")

                df_in_by_day = pd.read_sql("""
                    SELECT DATE(i.invoice_date) AS Дата, SUM(ii.qty) AS Приход
                    FROM invoice_items ii
                    JOIN invoices i ON i.id = ii.invoice_id
                    WHERE ii.product_id = ?
                      AND DATE(i.invoice_date) <= DATE(?)
                    GROUP BY DATE(i.invoice_date)
                """, conn, params=(selected_pid, str(selected_date)))

                df_out_by_day = pd.read_sql("""
                    SELECT DATE(s.sale_date) AS Дата, SUM(si.qty) AS Расход
                    FROM sale_items si
                    JOIN sales s ON s.id = si.sale_id
                    WHERE si.product_id = ?
                      AND DATE(s.sale_date) <= DATE(?)
                    GROUP BY DATE(s.sale_date)
                """, conn, params=(selected_pid, str(selected_date)))

                if df_in_by_day.empty and df_out_by_day.empty:
                    st.info("Нет движений по этому товару до выбранной даты.")
                else:
                    df_move = pd.merge(df_in_by_day, df_out_by_day, on="Дата", how="outer").fillna(0)
                    df_move["Дата"] = pd.to_datetime(df_move["Дата"])
                    df_move = df_move.sort_values("Дата")
                    df_move["Сальдо"] = df_move["Приход"] - df_move["Расход"]
                    df_move["Остаток_накоп"] = df_move["Сальдо"].cumsum()

                    # Найдём первый день, когда ушли в минус
                    neg = df_move[df_move["Остаток_накоп"] < 0]
                    if not neg.empty:
                        first_bad = neg.iloc[0]
                        st.error(f"⚠️ Уход в минус впервые: {first_bad['Дата'].date()} (накоп. остаток {first_bad['Остаток_накоп']:.2f})")

                    df_move_show = df_move.copy()
                    df_move_show["Дата"] = df_move_show["Дата"].dt.date
                    st.dataframe(
                        df_move_show[["Дата", "Приход", "Расход", "Сальдо", "Остаток_накоп"]],
                        use_container_width=True,
                        hide_index=True
                    )

        else:
            st.caption("💡 Кликни по строке товара в таблице — открою продажи/приходы и движение по датам.")
# ---------------------------------------------------
# 🛒 5. Продажи
# ---------------------------------------------------
if menu == "🛒 Продажи":
    st.subheader("🛒 Учет продаж (чек)")

    col1, col2 = st.columns(2)
    with col1:
        sale_number = st.text_input("Номер чека *")
    with col2:
        sale_date = st.date_input("Дата продажи", value=date.today())
    comment = st.text_input("Комментарий", "")

    if st.button("Создать чек"):
        if not sale_number:
            st.error("Введите номер чека.")
        else:
            sale_id = insert_sale(sale_number, str(sale_date), comment)
            st.success(f"Чек №{sale_number} создан (ID={sale_id})")
            st.rerun()

    st.markdown("---")
    st.subheader("Добавление товаров в чек")

    sales_df = get_sales_df()
    if sales_df.empty:
        st.info("Чеков пока нет. Сначала создайте чек.")
    else:
        label = sales_df.apply(lambda r: f"{r['sale_number']} — {r['sale_date']}", axis=1)
        selected = st.selectbox("Выберите чек", label)
        sale_id = int(sales_df.loc[label == selected, "id"].iloc[0])

        search = st.text_input("🔍 Поиск товара (артикул или название)")
        products_df = get_products(search)

        if products_df.empty:
            st.warning("Не найдено товаров.")
        else:
            product_label = products_df.apply(lambda r: f"{r['name']} ({r['article']})", axis=1)
            sel = st.selectbox("Выберите товар", product_label)
            product_id = int(products_df.loc[product_label == sel, "id"].iloc[0])

            col1, col2, col3 = st.columns(3)
            with col1:
                qty = st.number_input("Количество", min_value=0.0, step=1.0)
            with col2:
                price = st.number_input("Цена продажи", min_value=0.0, step=0.01)
            with col3:
                if st.button("Добавить в чек"):
                    add_sale_item(sale_id, product_id, qty, price)
                    st.success("Товар добавлен в чек.")
                    st.rerun()

        items_df = get_sale_items_df(sale_id)
        if not items_df.empty:
            st.dataframe(items_df, use_container_width=True)
            q, s = sale_totals(sale_id)
            st.markdown(f"**Итого:** Кол-во: {q:.2f}, Сумма: {s:.2f}")

# ---------------------------------------------------
# 📊 6. Отчёты
# ---------------------------------------------------
if menu == "📊 Отчёты":
    st.subheader("📊 Отчёты")

    report_type = st.radio("Выберите отчёт:", [
    "Остатки на дату",
    "Продажи за период",
    "Книга учёта товаров",
    "Интернет-продажи (маркетплейс)"
    ])

    # -------------------------------
    # 🔁 Массовый пересчёт FIFO
    # -------------------------------
    st.markdown("---")
    st.markdown("### 🔁 Массовый пересчёт FIFO")
    st.caption(f"Пересчитать FIFO для всех чеков, начиная с {FIFO_START_DATE}")
    if st.button("🔁 Пересчитать FIFO для всех чеков (начиная с FIFO_START_DATE)"):
        errors = []
        try:
            logger.info("Starting mass FIFO recalculation from %s", FIFO_START_DATE)
        except Exception:
            pass
        rows = c.execute("SELECT id, sale_number FROM sales WHERE DATE(sale_date) >= ? ORDER BY sale_date", (str(FIFO_START_DATE),)).fetchall()
        total = len(rows)
        if total == 0:
            st.info("Нет чеков для пересчёта в выбранном периоде.")
        else:
            prog = st.progress(0)
            log_box = st.empty()
            i = 0
            for sid, sn in rows:
                i += 1
                try:
                    recalc_fifo_for_sale(conn, sid)
                except Exception as e:
                    errors.append((sid, sn, str(e)))
                    try:
                        logger.exception("Error recalculating FIFO for sale_id=%s sale_number=%s: %s", sid, sn, e)
                    except Exception:
                        pass
                    # записываем в лог видимо для пользователя
                    log_box.text(f"Ошибка по чеку {sn} (id={sid}): {e}")
                prog.progress(int(i / total * 100))
            prog.progress(100)
            if errors:
                try:
                    logger.warning("Mass FIFO recalculation finished with %s errors", len(errors))
                except Exception:
                    pass
                st.error(f"Пересчёт завершён с {len(errors)} ошибками. См. лог ниже.")
                st.write(errors)
            else:
                try:
                    logger.info("Mass FIFO recalculation finished successfully for %s sales", total)
                except Exception:
                    pass
                st.success(f"Массовый пересчёт FIFO выполнен для {total} чеков.")
    st.markdown("---")

    # -------------------------------
    # 🔹 ОСТАТКИ НА ДАТУ
    # -------------------------------
    if report_type == "Остатки на дату":
        st.subheader("📦 Отчёт: Остатки товаров на дату")

        selected_date = st.date_input("Выберите дату", value=date.today())
        if st.button("📋 Показать отчёт"):
            df = get_stock_on_date(selected_date)

            if df.empty:
                st.info("Нет данных по остаткам на выбранную дату.")
            else:
                st.dataframe(df, use_container_width=True)
                total_sum = df["Сумма"].sum()
                st.markdown(f"**Итого по остаткам:** {total_sum:,.2f} BYN")

                buffer = export_stock_to_excel(df, selected_date)
                st.download_button(
                    label="💾 Экспорт в Excel",
                    data=buffer,
                    file_name=f"Остатки_{selected_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # -------------------------------
    # 🔹 ПРОДАЖИ ЗА ПЕРИОД
    # -------------------------------
    if report_type == "Продажи за период":
        st.subheader("📊 Отчёт: Продажи за период")

        col1, col2 = st.columns(2)
        with col1:
            date_from = st.date_input("Дата с")
        with col2:
            date_to = st.date_input("Дата по")

        if st.button("📋 Показать продажи"):
            df_sales = pd.read_sql("""
    SELECT 
        s.id AS sale_id,
        s.sale_number,
        s.sale_date,
        p.id AS product_id,
        p.article,
        p.name,
        p.unit,
        si.qty,
        si.price,
        si.total,
        COALESCE(si.gross_total, si.total) AS gross_total,
        COALESCE(si.mp_fee, 0) AS mp_fee,
        COALESCE(si.mp_delivery, 0) AS mp_delivery,
        COALESCE(si.net_total, si.total) AS net_total
    FROM sale_items si
    JOIN sales s ON s.id = si.sale_id
    JOIN products p ON p.id = si.product_id
    WHERE s.sale_date BETWEEN ? AND ?
    ORDER BY s.sale_date, s.sale_number
""", conn, params=(str(date_from), str(date_to)))

            if df_sales.empty:
                st.info("Нет продаж за этот период.")
            else:
                # Агрегированные показатели
                # Агрегированные показатели
                gross_sum = df_sales["gross_total"].sum() if "gross_total" in df_sales.columns else df_sales["total"].sum()
                fee_sum = df_sales["mp_fee"].sum() if "mp_fee" in df_sales.columns else 0.0
                delivery_sum = df_sales["mp_delivery"].sum() if "mp_delivery" in df_sales.columns else 0.0
                net_sum = df_sales["net_total"].sum() if "net_total" in df_sales.columns else (gross_sum - fee_sum - delivery_sum)

                total_qty = df_sales["qty"].sum()
                unique_sku = df_sales["name"].nunique()
                checks_count = df_sales["sale_number"].nunique()
                avg_check = df_sales.groupby("sale_number")["gross_total"].sum().mean() if "gross_total" in df_sales.columns else df_sales.groupby("sale_number")["total"].sum().mean()


                st.markdown("### 📌 Итоги периода")
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("Чеков", checks_count)
                c2.metric("Выручка (грязная)", f"{gross_sum:.2f}")
                c3.metric("Средний чек", f"{avg_check:.2f}")
                c4.metric("Продано единиц", f"{total_qty:.2f}")
                c5.metric("Уникальных товаров", unique_sku)

                # Для маркетплейса покажем удержания и чистую выручку
                d1, d2, d3 = st.columns(3)
                d1.metric("Комиссия", f"{fee_sum:.2f}")
                d2.metric("Доставка", f"{delivery_sum:.2f}")
                d3.metric("Чистая выручка", f"{net_sum:.2f}")

                st.markdown("---")

                # TOP-10 товаров
                st.markdown("### 🏆 TOP-10 товаров")
                top_products = (
                    df_sales.groupby("name")["gross_total"]
                    .sum()
                    .sort_values(ascending=False)
                    .head(10)
                )
                st.bar_chart(top_products)

                st.markdown("---")

                # Продажи по дням
                st.markdown("### 📅 Продажи по дням")
                df_by_day = df_sales.groupby("sale_date")["gross_total"].sum().reset_index()
                st.line_chart(df_by_day, x="sale_date", y="gross_total")

                st.markdown("---")

                # Детальная таблица
                st.markdown("### 📋 Детализация продаж (агрегировано по товарам)")

                df_grouped = (
                    df_sales.groupby(["article", "name", "unit"])
                    .agg({
                        "qty": "sum",
                        "price": "mean",
                        "gross_total": "sum",
                        "mp_fee": "sum",
                        "mp_delivery": "sum",
                        "net_total": "sum"
                    })
                    .reset_index()
                )

                df_grouped = df_grouped.rename(columns={
                    "article": "Артикул",
                    "name": "Наименование",
                    "unit": "Ед",
                    "qty": "Количество",
                    "price": "Средняя цена",
                    "gross_total": "Выручка",
                    "mp_fee": "Комиссия",
                    "mp_delivery": "Доставка",
                    "net_total": "Чистая выручка"
                })

                # СОРТИРОВКА ПО КОЛИЧЕСТВУ (по убыванию) 
                df_grouped = df_grouped.sort_values(by="Количество", ascending=False)

                st.dataframe(df_grouped, use_container_width=True)

                
                # Экспорт в Excel (агрегированной таблицы!)
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
                    df_grouped.to_excel(writer, index=False, sheet_name="Агрегат")

                buffer.seek(0)

                st.download_button(
                    label="💾 Скачать Excel (агрегировано)",
                    data=buffer,
                    file_name=f"Продажи_агрегат_{date_from}_to_{date_to}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            if not df_sales.empty:
                # Загружаем все приходы до date_to
                df_in = pd.read_sql("""
                    SELECT 
                        ii.id AS item_id,
                        ii.product_id,
                        ii.qty AS qty_in,
                        ii.price AS price_in,
                        i.invoice_date
                    FROM invoice_items ii
                    JOIN invoices i ON i.id = ii.invoice_id
                """, conn)

                df_in["invoice_date"] = pd.to_datetime(df_in["invoice_date"])
                df_in = df_in[df_in["invoice_date"] <= pd.to_datetime(date_to)]
                df_in = df_in.sort_values(["product_id", "invoice_date", "item_id"])

                # Подготовка FIFO структур
                fifo_batches = {}
                for _, r in df_in.iterrows():
                    pid = int(r["product_id"])
                    if pid not in fifo_batches:
                        fifo_batches[pid] = []
                    fifo_batches[pid].append({
                        "qty": float(r["qty_in"]),
                        "price": float(r["price_in"]),
                    })

                total_cogs = 0.0  # себестоимость всего периода

                # Сортируем продажи по дате
                df_sales["sale_date"] = pd.to_datetime(df_sales["sale_date"])
                df_sales = df_sales.sort_values(["product_id", "sale_date"])

                # FIFO списание
                for _, s in df_sales.iterrows():
                    pid = int(s["product_id"])
                    qty_sold = float(s["qty"])
                    if pid not in fifo_batches:
                        continue  # нет приходов → пропускаем

                    batches = fifo_batches[pid]

                    i = 0
                    while qty_sold > 0 and i < len(batches):
                        batch = batches[i]
                        if batch["qty"] <= 0:
                            i += 1
                            continue

                        allocated = min(qty_sold, batch["qty"])

                        # COGS по этой партии
                        total_cogs += allocated * batch["price"]

                        # уменьшаем остатки
                        batch["qty"] -= allocated
                        qty_sold -= allocated

                # Добавляем в интерфейс
                st.metric("Себестоимость (COGS)", f"{total_cogs:.2f}")

                # Прибыль для налогов: грязная выручка - комиссия - доставка - себестоимость
                profit_tax = gross_sum - fee_sum - delivery_sum - total_cogs
                st.metric("Прибыль", f"{profit_tax:.2f}")

                st.markdown("### 🧾 Налоговый расчёт за период")
                tax_base = profit_tax  # по вашей формуле налоговая база = прибыль для налогов

                t1, t2, t3, t4, t5 = st.columns(5)
                t1.metric("Грязная выручка", f"{gross_sum:,.2f}")
                t2.metric("Комиссия", f"{fee_sum:,.2f}")
                t3.metric("Доставка", f"{delivery_sum:,.2f}")
                t4.metric("Себестоимость (FIFO)", f"{total_cogs:,.2f}")
                t5.metric("НАЛОГОВАЯ БАЗА", f"{tax_base:,.2f}")

                # Экспорт налогового расчёта в Excel
                tax_df = pd.DataFrame([
                    {"Показатель": "Период с", "Сумма": str(date_from)},
                    {"Показатель": "Период по", "Сумма": str(date_to)},
                    {"Показатель": "Грязная выручка", "Сумма": round(float(gross_sum), 2)},
                    {"Показатель": "Комиссия", "Сумма": round(float(fee_sum), 2)},
                    {"Показатель": "Доставка", "Сумма": round(float(delivery_sum), 2)},
                    {"Показатель": "Себестоимость (FIFO)", "Сумма": round(float(total_cogs), 2)},
                    {"Показатель": "Налоговая база", "Сумма": round(float(tax_base), 2)},
                ])

                tax_buf = BytesIO()
                with pd.ExcelWriter(tax_buf, engine="xlsxwriter") as writer:
                    tax_df.to_excel(writer, index=False, sheet_name="Налоги")
                tax_buf.seek(0)

                st.download_button(
                    label="💾 Экспорт налогового расчёта в Excel",
                    data=tax_buf,
                    file_name=f"Налоговый_расчет_{date_from}_to_{date_to}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # -------------------------------
    # 🔹 КНИГА УЧЁТА ТОВАРОВ
    # -------------------------------
    if report_type == "Книга учёта товаров":
        st.subheader("📘 Выгрузка: Книга учёта товаров")

        selected_year = st.number_input("Год", min_value=2020, max_value=2100, value=date.today().year)

        if st.button("📤 Сформировать Excel"):
            buffer = export_book(selected_year)

            if buffer is None:
                st.warning("Нет данных за этот год.")
            else:
                st.success("Файл успешно сформирован!")
                st.download_button(
                    label="💾 Скачать файл",
                    data=buffer,
                    file_name=f"Книга_учёта_товаров_{selected_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    # -------------------------------
    # 🔹 ИНТЕРНЕТ-ПРОДАЖИ (МАРКЕТПЛЕЙС)
    # -------------------------------
    
    if report_type == "Интернет-продажи (маркетплейс)":
        st.subheader("🌐 Отчёт: Интернет-продажи маркетплейса")

        st.markdown("### Шаг 1 — Загрузить файлы и сформировать черновик")
        wb_file = st.file_uploader("1) Файл продаж (Детализация оказанных услуг)", type=["xlsx"], key="mp_wb_file")
        ep_files = st.file_uploader(
            "2) Файлы Европочты (доставка) — можно несколько",
            type=["xlsx"],
            accept_multiple_files=True,
            key="mp_ep_files"
        )

        colA, colB, colC = st.columns([1, 1, 1])
        with colA:
            if st.button("🧾 Сформировать черновик месячного чека", disabled=not bool(wb_file)):
                try:
                    merged_df, stats, missing_delivery_df = build_internet_sales_from_files(wb_file, ep_files)

                    # Покажем сводку по доставке
                    st.session_state.mp_draft_stats = stats or {}
                    st.session_state.mp_draft_missing_delivery_df = missing_delivery_df

                    # Формируем строки месячного чека (защита от дублей):
                    # 1) схлопываем одинаковые "Артикул" (фактически ШК) в одну строку
                    # 2) далее сопоставляем с товарами в базе
                    month_df = merged_df.copy()
                    month_df["Артикул"] = month_df["Артикул"].astype(str).str.strip()
                    month_df["Количество"] = pd.to_numeric(month_df["Количество"], errors="coerce").fillna(0.0)
                    month_df["Цена продажи, руб."] = pd.to_numeric(month_df["Цена продажи, руб."], errors="coerce").fillna(0.0)

                    # Сумма по строке
                    if "Сумма" not in month_df.columns:
                        month_df["Сумма"] = (month_df["Количество"] * month_df["Цена продажи, руб."]).round(2)
                    else:
                        month_df["Сумма"] = pd.to_numeric(month_df["Сумма"], errors="coerce").fillna(0.0)

                    
                    # Защита от дублей: по одному SKU на месяц.
                    # ВАЖНО: сохраняем комиссию/доставку/чистую выручку, рассчитанные в build_internet_sales_from_files
                    # (они уже распределены по строкам заказа).
                    if "Комиссия" not in month_df.columns:
                        month_df["Комиссия"] = 0.0
                    if "Доставка" not in month_df.columns:
                        month_df["Доставка"] = 0.0
                    if "Чистая_сумма" not in month_df.columns:
                        month_df["Чистая_сумма"] = (month_df["Сумма"] - month_df["Комиссия"] - month_df["Доставка"]).round(2)

                    month_agg = (
                        month_df
                        .groupby(["Артикул"], as_index=False)
                        .agg({
                            "Количество": "sum",
                            "Сумма": "sum",
                            "Комиссия": "sum",
                            "Доставка": "sum",
                            "Чистая_сумма": "sum",
                            "Наименование товара": "first"
                        })
                    )
                    month_agg["Цена продажи, руб."] = (month_agg["Сумма"] / month_agg["Количество"]).replace([float("inf"), -float("inf")], 0).fillna(0).round(2)
                    month_agg["Чистая цена, руб."] = (month_agg["Чистая_сумма"] / month_agg["Количество"]).replace([float("inf"), -float("inf")], 0).fillna(0).round(2)
                    # Сопоставляем с базой: если заполнен products.barcode — матчим по нему,
                    # иначе пробуем по products.article (fallback).
                    prod_map = pd.read_sql(
                    "SELECT id AS product_id, article, name, COALESCE(barcode, '') AS barcode FROM products",
                    conn
                    )
                    prod_map["barcode"] = prod_map["barcode"].astype(str).str.strip()
                    prod_map["article"] = prod_map["article"].astype(str).str.strip()

                    # 1) матч по barcode
                    month_lines = month_agg.merge(prod_map, how="left", left_on="Артикул", right_on="barcode")

                    # 2) fallback по article для тех, кто не совпал по barcode
                    miss_mask = month_lines["product_id"].isna()
                    if miss_mask.any():
                        month_lines_fb = month_agg.merge(prod_map, how="left", left_on="Артикул", right_on="article")
                        for col in ["product_id", "article", "name", "barcode"]:
                            month_lines.loc[miss_mask, col] = month_lines_fb.loc[miss_mask, col].values

                    missing_products_df = month_lines[month_lines["product_id"].isna()].copy()
                    ok_lines = month_lines[month_lines["product_id"].notna()].copy()

                    # Готовим финальные строки для загрузки в sale_items
                    lines_df = ok_lines[["product_id", "article", "name", "Количество", "Цена продажи, руб.", "Сумма", "Комиссия", "Доставка", "Чистая_сумма"]].copy()
                    lines_df = lines_df.rename(columns={
                    "article": "Артикул",
                    "name": "Наименование",
                    "Цена продажи, руб.": "Цена",
                    })

                    st.session_state.mp_draft_lines_df = lines_df

                    st.session_state.mp_draft_missing_products_df = missing_products_df[["Артикул", "Наименование товара", "Количество", "Цена продажи, руб.", "Сумма"]].rename(columns={
                        "Артикул": "ШК",
                        "Наименование товара": "Наименование",
                        "Количество": "Количество",
                        "Цена продажи, руб.": "Цена",
                        "Сумма": "Сумма"
                    })
                    st.session_state.mp_draft_ready = True

                    st.success("Черновик сформирован. Перейди к шагу 2 ниже: задай номер и дату чека и загрузи в базу.")
                except Exception as e:
                    st.session_state.mp_draft_ready = False
                    st.error(f"Ошибка при формировании черновика: {e}")

        with colB:
            if st.button("🧹 Сбросить черновик"):
                st.session_state.mp_draft_ready = False
                st.session_state.mp_draft_lines_df = None
                st.session_state.mp_draft_missing_products_df = None
                st.session_state.mp_draft_stats = {}
                st.session_state.mp_draft_missing_delivery_df = None
                st.success("Черновик сброшен.")

        with colC:
            st.checkbox("Перезаписывать чек, если номер уже существует", key="mp_overwrite")

        # Показ статистики доставки (если есть)
        if st.session_state.mp_draft_stats:
            s = st.session_state.mp_draft_stats
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("Заказов в файле продаж", int(s.get("total_orders", 0)))
            c2.metric("Нашлось доставок", int(s.get("matched_orders", 0)))
            c3.metric("Не найдено доставок", int(s.get("missing_orders", 0)))
            c4.metric("Доставка в файлах, руб.", round(float(s.get("delivery_total_files", 0.0)), 2))
            c5.metric("Доставка применена, руб.", round(float(s.get("delivery_total_applied", 0.0)), 2))

        if st.session_state.mp_draft_missing_delivery_df is not None and not st.session_state.mp_draft_missing_delivery_df.empty:
            st.markdown("**Заказы без найденной стоимости доставки:**")
            st.dataframe(st.session_state.mp_draft_missing_delivery_df, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("### Шаг 2 — Проверить состав чека, указать номер/дату и загрузить в базу")

        if not st.session_state.mp_draft_ready or st.session_state.mp_draft_lines_df is None:
            st.info("Сначала нажми «Сформировать черновик месячного чека» (шаг 1).")
        else:
            # Показываем итоговые строки для sale_items
            st.markdown("**Состав месячного чека (уникальные товары, дубли схлопнуты):**")
            edited = st.data_editor(
                st.session_state.mp_draft_lines_df,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed"
            )
            st.session_state.mp_draft_lines_df = edited

            total_sum = float(pd.to_numeric(edited["Сумма"], errors="coerce").fillna(0).sum())
            m1, m2 = st.columns(2)
            m1.metric("Позиций (уникальных товаров)", int(len(edited)))
            m2.metric("Сумма продаж за месяц", round(total_sum, 2))

            # Ошибки по товарам (неверный ШК/нет в базе)
            missing_products_df = st.session_state.mp_draft_missing_products_df
            if missing_products_df is not None and not missing_products_df.empty:
                st.warning("Есть товары, которых нет в базе (скорее всего неверный ШК/артикул в файле). Эти позиции НЕ будут загружены.")
                st.dataframe(missing_products_df, use_container_width=True, hide_index=True)

                # Скачать ошибки в Excel
                out = BytesIO()
                with pd.ExcelWriter(out, engine="openpyxl") as writer:
                    missing_products_df.to_excel(writer, index=False, sheet_name="MissingProducts")
                st.download_button(
                    "⬇️ Скачать ошибки (не найден товар) в Excel",
                    data=out.getvalue(),
                    file_name="missing_products.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            # Форма для номера/даты чека — отдельно, стабильно
            with st.form("mp_finalize_form", clear_on_submit=False):
                st.text_input("Номер чека (как ты хочешь видеть в базе)", key="mp_sale_number", placeholder="например: MP-04-2025")
                st.date_input("Дата чека", key="mp_sale_date")
                submit = st.form_submit_button("✅ Подтвердить номер и дату")

            st.info(f"Будет создан чек: **{st.session_state.mp_sale_number or '(не задан)'}** от **{st.session_state.mp_sale_date}**")

            can_load = bool(st.session_state.mp_sale_number.strip()) and edited is not None and not edited.empty and (missing_products_df is None or missing_products_df.empty)

            if st.button("⬇️ Загрузить в базу (1 чек за месяц)", disabled=not can_load):
                res = push_marketplace_month_to_db(
                    st.session_state.mp_draft_lines_df,
                    st.session_state.mp_sale_number.strip(),
                    st.session_state.mp_sale_date,
                    overwrite=bool(st.session_state.mp_overwrite)
                )
                if not res.get("ok"):
                    st.error(res.get("error", "Ошибка загрузки."))
                else:
                    st.success(f"Готово! В базе создан/обновлён чек **{st.session_state.mp_sale_number.strip()}** от **{st.session_state.mp_sale_date}**. Для поиска: 'Просмотр базы' → 'Продажи (чеки)'.")
                    st.rerun()

# =======================
# v12 – Усиленный контроль остатков + таймлайн
# =======================

from datetime import timedelta

def get_stock_on_date(product_id: int, on_date: str) -> float:
    in_qty = c.execute("""
        SELECT COALESCE(SUM(ii.qty), 0)
        FROM invoice_items ii
        JOIN invoices i ON i.id = ii.invoice_id
        WHERE ii.product_id = ?
          AND DATE(i.invoice_date) <= DATE(?)
    """, (product_id, on_date)).fetchone()[0]

    out_qty = c.execute("""
        SELECT COALESCE(SUM(si.qty), 0)
        FROM sale_items si
        JOIN sales s ON s.id = si.sale_id
        WHERE si.product_id = ?
          AND DATE(s.sale_date) <= DATE(?)
    """, (product_id, on_date)).fetchone()[0]

    return float(in_qty - out_qty)


def get_available_stock(product_id: int, on_date: str, exclude_sale_id: int | None = None) -> float:
    in_qty = c.execute("""
        SELECT COALESCE(SUM(ii.qty), 0)
        FROM invoice_items ii
        JOIN invoices i ON i.id = ii.invoice_id
        WHERE ii.product_id = ?
          AND DATE(i.invoice_date) <= DATE(?)
    """, (product_id, on_date)).fetchone()[0]

    if exclude_sale_id:
        out_qty = c.execute("""
            SELECT COALESCE(SUM(si.qty), 0)
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE si.product_id = ?
              AND DATE(s.sale_date) <= DATE(?)
              AND s.id != ?
        """, (product_id, on_date, exclude_sale_id)).fetchone()[0]
    else:
        out_qty = c.execute("""
            SELECT COALESCE(SUM(si.qty), 0)
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE si.product_id = ?
              AND DATE(s.sale_date) <= DATE(?)
        """, (product_id, on_date)).fetchone()[0]

    return float(in_qty - out_qty)


def safe_add_sale_item(sale_id, product_id, qty, price):
    sale_date = c.execute(
        "SELECT sale_date FROM sales WHERE id = ?",
        (sale_id,)
    ).fetchone()[0]

    available = get_available_stock(
        product_id=product_id,
        on_date=sale_date,
        exclude_sale_id=sale_id
    )

    if qty > available:
        raise ValueError(
            f"Недостаточно остатка. Доступно: {available:.2f}, попытка: {qty:.2f}"
        )

    add_sale_item(sale_id, product_id, qty, price)


def get_product_timeline(product_id: int):
    df = pd.read_sql("""
        SELECT
            i.invoice_date AS date,
            'Приход' AS type,
            i.id AS doc_id,
            i.invoice_number AS doc_number,
            ii.qty AS qty
        FROM invoice_items ii
        JOIN invoices i ON i.id = ii.invoice_id
        WHERE ii.product_id = ?

        UNION ALL

        SELECT
            s.sale_date AS date,
            'Продажа' AS type,
            s.id AS doc_id,
            s.sale_number AS doc_number,
            -si.qty AS qty
        FROM sale_items si
        JOIN sales s ON s.id = si.sale_id
        WHERE si.product_id = ?

        ORDER BY date
    """, conn, params=(product_id, product_id))

    if df.empty:
        return df

    df["date"] = pd.to_datetime(df["date"])
    df["qty"] = df["qty"].astype(float)
    df["balance"] = df["qty"].cumsum()

    return df

# =======================
# Конец v12
# =======================


# ================= FIFO v15 =================
from datetime import date

FIFO_START_DATE = date(2026, 1, 1)

def ensure_sale_fifo(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sale_fifo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_item_id INTEGER NOT NULL,
            batch_id INTEGER NOT NULL,
            qty REAL NOT NULL,
            cost_price REAL NOT NULL
        )
    """)
    conn.commit()
    try:
        logger.info("Ensured sale_fifo table exists")
    except Exception:
        pass

def recalc_fifo_for_sale(conn, sale_id):
    try:
        logger.info("Recalculating FIFO for sale_id=%s", sale_id)
    except Exception:
        pass

    # delete old fifo for this sale
    conn.execute("""
        DELETE FROM sale_fifo
        WHERE sale_item_id IN (
            SELECT id FROM sale_items WHERE sale_id = ?
        )
    """, (sale_id,))
    conn.commit()

    items = conn.execute(
        "SELECT id, product_id, qty FROM sale_items WHERE sale_id = ?",
        (sale_id,)
    ).fetchall()

    sale_date = conn.execute(
        "SELECT sale_date FROM sales WHERE id = ?",
        (sale_id,)
    ).fetchone()[0]

    for si_id, product_id, qty_need in items:
        try:
            logger.debug("Allocating sale_item_id=%s product_id=%s qty_need=%s", si_id, product_id, qty_need)
        except Exception:
            pass
        batches = conn.execute("""
            SELECT ii.id, ii.qty, ii.price
            FROM invoice_items ii
            JOIN invoices i ON i.id = ii.invoice_id
            WHERE ii.product_id = ? AND i.invoice_date <= ?
            ORDER BY i.invoice_date, ii.id
        """, (product_id, sale_date)).fetchall()

        remaining = qty_need
        for batch_id, batch_qty, cost in batches:
            used = conn.execute(
                "SELECT COALESCE(SUM(qty),0) FROM sale_fifo WHERE batch_id = ?",
                (batch_id,)
            ).fetchone()[0]
            available = batch_qty - used
            if available <= 0:
                continue
            take = min(available, remaining)
            conn.execute(
                "INSERT INTO sale_fifo (sale_item_id, batch_id, qty, cost_price) VALUES (?,?,?,?)",
                (si_id, batch_id, take, cost)
            )
            try:
                logger.debug("Inserted FIFO: sale_item=%s batch=%s qty=%s cost=%s", si_id, batch_id, take, cost)
            except Exception:
                pass
            remaining -= take
            if remaining <= 0:
                break
    conn.commit()
    try:
        logger.info("Finished FIFO recalculation for sale_id=%s", sale_id)
    except Exception:
        pass
# ================= END FIFO v15 =================


# ================= FIFO UI v16 =================

def show_fifo_for_sale(conn, sale_id):
    fifo = conn.execute("""
        SELECT
            sf.qty,
            sf.cost_price,
            ii.id AS batch_id,
            i.invoice_number,
            i.invoice_date
        FROM sale_fifo sf
        JOIN invoice_items ii ON ii.id = sf.batch_id
        JOIN invoices i ON i.id = ii.invoice_id
        JOIN sale_items si ON si.id = sf.sale_item_id
        WHERE si.sale_id = ?
        ORDER BY i.invoice_date
    """, (sale_id,)).fetchall()

    st.markdown("### 📦 FIFO-списание")

    if not fifo:
        st.info("FIFO ещё не рассчитан")
        return

    total = 0
    for qty, price, batch_id, inv_num, inv_date in fifo:
        amount = qty * price
        total += amount
        c1, c2, c3, c4, c5 = st.columns([2,2,1,1,1])
        c1.write(inv_num)
        c2.write(inv_date)
        c3.write(qty)
        c4.write(price)
        if c5.button("➡", key=f"batch_{batch_id}"):
            st.session_state.nav = {
                "section": "📦 Склад",
                "batch_id": batch_id,
                "sale_id": sale_id
            }
            st.rerun()

    st.metric("Себестоимость чека", f"{total:.2f}")

# ================= END FIFO UI v16 =================
